from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from sqlalchemy.orm import close_all_sessions

from app import db
from app.domain import models
from app.domain.models import Base
from app.schemas import (
    AnalysisBatchCreateInput,
    AnalysisBatchItemInput,
    AnalysisImportCommitInput,
    BatchModifyImportCommitInput,
    AnalysisPreviewRequest,
    BoxCreateInput,
    BulkBoxImportCommitInput,
    BulkSampleImportCommitInput,
    EventSearchQuery,
    PlaceSampleInput,
    RetrieveSampleInput,
    SampleCreateInput,
    SampleNoteCreateInput,
    SampleSearchQuery,
    SampleUpdateInput,
    SampleTypeCreate,
    StorageNodeCreate,
    StorageNodeMoveInput,
    StorageNodeUpdate,
    StudyCreate,
)
from app.services import admin as admin_service
from app.services import analyses as analysis_service
from app.services import auth as auth_service
from app.services import events as event_service
from app.services import batch_modify as batch_modify_service
from app.services import bulk_imports as bulk_import_service
from app.services import samples as sample_service
from app.services import storage as storage_service


class SampleWorkflowServiceTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        db_path = Path(self.tempdir.name, "service-tests.db").resolve().as_posix()
        db.configure(f"sqlite:///{db_path}")
        Base.metadata.drop_all(bind=db.engine)
        Base.metadata.create_all(bind=db.engine)
        self.session = db.SessionLocal()
        self.user = auth_service.sync_user(self.session, "admin", "Admin")
        self.study = admin_service.create_study(self.session, StudyCreate(name="Sovary"))
        self.sample_type = admin_service.create_sample_type(
            self.session,
            SampleTypeCreate(name="Plasma", description="Plasma"),
        )

        self.freezer = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Freezer A", notes="Primary freezer", node_type="freezer"),
            self.user,
        )
        self.shelf = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Shelf 1", node_type="shelf", parent_id=self.freezer.id),
            self.user,
        )
        self.rack = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Rack 1", node_type="rack", parent_id=self.shelf.id),
            self.user,
        )
        self.box = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Box 1", node_type="box", parent_id=self.rack.id),
            self.user,
        )
        storage_service.create_box_positions(
            self.session,
            BoxCreateInput(box_id=self.box.id, rows=2, cols=2),
            self.user,
        )
        self.box_two = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Box 2", node_type="box", parent_id=self.rack.id),
            self.user,
        )
        storage_service.create_box_positions(
            self.session,
            BoxCreateInput(box_id=self.box_two.id, rows=2, cols=2),
            self.user,
        )
        self.positions = storage_service.get_box_view(self.session, self.box.id).positions
        self.box_two_positions = storage_service.get_box_view(self.session, self.box_two.id).positions

    def tearDown(self):
        self.session.close()
        close_all_sessions()
        db.engine.dispose()
        self.tempdir.cleanup()

    def create_and_place_sample(
        self,
        sample_identifier: str,
        *,
        position_id: int,
        volume: float = 1.0,
        study_role: str = "current",
    ):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id=sample_identifier,
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                volume=volume,
                study_role=study_role,
            ),
            self.user,
        )
        sample_service.place_sample(
            self.session,
            sample.id,
            PlaceSampleInput(position_id=position_id),
            self.user,
        )
        return sample

    def test_sample_metadata_defaults_and_note_log(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-001",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="60",
                aliquot_number=1,
                notes="Baseline aliquot",
                collection_at=datetime(2026, 3, 9, 8, 30),
            ),
            self.user,
        )
        detail = sample_service.get_sample_detail(self.session, sample.id)
        self.assertEqual(detail.study_role, "current")
        self.assertEqual(detail.custody_label, "unplaced")
        self.assertEqual(detail.usage_label, "unused")
        self.assertEqual(detail.thaw_count, 0)
        self.assertIsNone(detail.hemolysis_classification)
        self.assertEqual(detail.study_name, "Sovary")
        self.assertEqual(detail.visit_label, "1")
        self.assertEqual(detail.timepoint_label, "60")

        sample_service.add_note_entry(
            self.session,
            sample.id,
            SampleNoteCreateInput(text="Used 0.2 mL for assay"),
            self.user,
        )
        detail = sample_service.get_sample_detail(self.session, sample.id)
        self.assertEqual(len(detail.note_entries), 1)
        self.assertEqual(detail.note_entries[0].text, "Used 0.2 mL for assay")

    def test_first_user_bootstraps_as_admin_and_later_users_default_to_staff(self):
        self.assertEqual(self.user.role.value, "admin")
        second_user = auth_service.sync_user(self.session, "second", "Second User")
        self.assertEqual(second_user.role.value, "staff")

    def test_admin_can_manage_user_roles_but_last_admin_must_remain(self):
        staff_user = auth_service.sync_user(self.session, "staffer", "Staff User")
        updated = auth_service.update_user_admin(
            self.session,
            self.user,
            staff_user.id,
            full_name="Staff Updated",
            role="admin",
            permissions=list(auth_service.default_permissions_for_role("admin")),
        )
        self.assertEqual(updated.role.value, "admin")
        self.assertEqual(updated.full_name, "Staff Updated")

        auth_service.update_user_admin(
            self.session,
            self.user,
            staff_user.id,
            full_name="Staff Updated",
            role="staff",
            permissions=list(auth_service.default_permissions_for_role("staff")),
        )

        with self.assertRaises(auth_service.UserManagementError):
            auth_service.update_user_admin(
                self.session,
                self.user,
                self.user.id,
                full_name="Admin",
                role="staff",
                permissions=list(auth_service.default_permissions_for_role("staff")),
            )

    def test_role_defaults_and_user_permission_overrides(self):
        staff_user = auth_service.sync_user(self.session, "ops", "Ops User")
        self.assertTrue(auth_service.has_permission(staff_user, "edit_samples"))
        self.assertFalse(auth_service.has_permission(staff_user, "manage_vocabularies"))

        updated = auth_service.update_user_admin(
            self.session,
            self.user,
            staff_user.id,
            full_name="Ops User",
            role="staff",
            permissions=[
                "edit_samples",
                "archive_samples",
                "place_move_samples",
                "bulk_import_samples",
                "manage_vocabularies",
            ],
        )
        self.assertTrue(auth_service.has_permission(updated, "manage_vocabularies"))
        self.assertFalse(auth_service.has_permission(updated, "manage_storage_tree"))

        limited_admin = auth_service.update_user_admin(
            self.session,
            self.user,
            staff_user.id,
            full_name="Ops User",
            role="admin",
            permissions=[
                "manage_users",
                "manage_vocabularies",
                "edit_samples",
            ],
        )
        self.assertTrue(auth_service.has_permission(limited_admin, "manage_users"))
        self.assertTrue(auth_service.has_permission(limited_admin, "manage_vocabularies"))
        self.assertFalse(auth_service.has_permission(limited_admin, "manage_storage_tree"))

    def test_same_participant_id_can_have_multiple_samples_but_not_duplicate_composite_identity(self):
        first = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="P-001",
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
            ),
            self.user,
        )
        self.assertIsNotNone(first)

        second = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="P-001",
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="60",
                aliquot_number=1,
            ),
            self.user,
        )
        self.assertIsNotNone(second)

        with self.assertRaises(sample_service.SampleError):
            sample_service.create_sample(
                self.session,
                SampleCreateInput(
                    sample_id="P-001",
                    sample_type_id=self.sample_type.id,
                    visit_label="1",
                    timepoint_label="00",
                    aliquot_number=1,
                ),
                self.user,
            )

    def test_admin_can_update_sample_type_and_study(self):
        updated_sample_type = admin_service.update_sample_type(
            self.session,
            self.sample_type.id,
            SampleTypeCreate(name="Serum", description="Updated description"),
        )
        self.assertEqual(updated_sample_type.name, "Serum")
        self.assertEqual(updated_sample_type.description, "Updated description")

        updated_study = admin_service.update_study(
            self.session,
            self.study.id,
            StudyCreate(name="Sovary Renamed", description="Updated study"),
        )
        self.assertEqual(updated_study.name, "Sovary Renamed")
        self.assertEqual(updated_study.description, "Updated study")
        self.assertEqual(updated_study.name, "Sovary Renamed")

    def test_admin_can_delete_unreferenced_sample_type_and_study(self):
        disposable_type = admin_service.create_sample_type(
            self.session,
            SampleTypeCreate(name="Disposable Type", description="To delete"),
        )
        disposable_study = admin_service.create_study(
            self.session,
            StudyCreate(name="Disposable Study", description="To delete"),
        )

        admin_service.delete_sample_type(self.session, disposable_type.id)
        admin_service.delete_study(self.session, disposable_study.id)

        self.assertFalse(any(item.id == disposable_type.id for item in admin_service.list_sample_types(self.session)))
        self.assertFalse(any(item.id == disposable_study.id for item in admin_service.list_studies(self.session)))

    def test_admin_cannot_delete_referenced_sample_type_or_study(self):
        sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="REF-001",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                collection_at=datetime(2026, 3, 10, 8, 0),
            ),
            self.user,
        )

        with self.assertRaises(admin_service.AdminError):
            admin_service.delete_sample_type(self.session, self.sample_type.id)

        with self.assertRaises(admin_service.AdminError):
            admin_service.delete_study(self.session, self.study.id)

    def test_place_and_archive_sample(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="S-010", study_id=self.study.id),
            self.user,
        )

        placement = sample_service.place_sample(
            self.session,
            sample.id,
            PlaceSampleInput(position_id=self.positions[0].id),
            self.user,
        )
        self.assertEqual(placement.position_id, self.positions[0].id)

        archived = sample_service.retrieve_sample(
            self.session,
            sample.id,
            RetrieveSampleInput(note="Shipment complete"),
            self.user,
        )
        self.assertTrue(archived.is_archived)
        self.assertFalse(archived.is_out_for_analysis)
        self.assertIsNone(sample_service.get_sample_detail(self.session, sample.id).location_label)

        with self.assertRaises(sample_service.SampleError):
            sample_service.place_sample(
                self.session,
                sample.id,
                PlaceSampleInput(position_id=self.positions[1].id),
                self.user,
            )

    def test_analysis_batch_returns_samples_and_records_audit(self):
        first = self.create_and_place_sample("AN-001", position_id=self.positions[0].id, volume=1.0)
        second = self.create_and_place_sample("AN-002", position_id=self.positions[1].id, volume=0.8)

        result = analysis_service.submit_analysis_batch(
            self.session,
            AnalysisBatchCreateInput(
                analysis_type="Chemistry Panel",
                performed_at=datetime(2026, 3, 27, 9, 30),
                overall_notes="Morning run",
                items=[
                    AnalysisBatchItemInput(sample_id=first.id, remaining_volume=0.6, thaw_increment=1, returned_to_storage=True),
                    AnalysisBatchItemInput(sample_id=second.id, remaining_volume=0.4, thaw_increment=2, returned_to_storage=True),
                ],
            ),
            self.user,
        )

        self.assertEqual(result.returned_count, 2)
        self.assertEqual(result.out_for_analysis_count, 0)
        self.assertEqual(result.processed_count, 2)

        first_detail = sample_service.get_sample_detail(self.session, first.id)
        second_detail = sample_service.get_sample_detail(self.session, second.id)
        self.assertEqual(first_detail.study_role, "current")
        self.assertEqual(first_detail.custody_label, "in storage")
        self.assertEqual(first_detail.usage_label, "used")
        self.assertEqual(first_detail.volume, 0.6)
        self.assertEqual(first_detail.thaw_count, 1)
        self.assertEqual(first_detail.location_label, "A1")
        self.assertEqual(second_detail.custody_label, "in storage")
        self.assertEqual(second_detail.usage_label, "used")
        self.assertEqual(second_detail.thaw_count, 2)
        self.assertTrue(first_detail.note_entries[0].text.startswith("Analysis: Chemistry Panel"))

        analysis_batches = self.session.query(models.AnalysisBatch).all()
        analysis_items = self.session.query(models.AnalysisItem).all()
        self.assertEqual(len(analysis_batches), 1)
        self.assertEqual(len(analysis_items), 2)
        self.assertTrue(all(item.batch_id == analysis_batches[0].id for item in analysis_items))

        analysis_events = event_service.list_events(
            self.session,
            query=EventSearchQuery(event_type="analyze_sample", sample_query="AN-001", limit=20),
        )
        self.assertEqual(len(analysis_events), 2)
        self.assertTrue(analysis_events[0].is_group_parent)
        self.assertEqual(analysis_events[0].group_kind, "analysis_batch")
        self.assertEqual(analysis_events[0].group_id, str(analysis_batches[0].id))
        self.assertEqual(analysis_events[1].action_label, "Analysis")
        self.assertTrue(analysis_events[1].is_group_child)
        self.assertEqual(analysis_events[1].payload["analysis_batch_id"], analysis_batches[0].id)
        self.assertTrue(any(change.label == "Volume" and change.after == "0.6 mL" for change in analysis_events[1].change_items))
        self.assertEqual(analysis_events[1].payload["batch_group_kind"], "analysis_batch")
        self.assertEqual(analysis_events[1].payload["batch_group_id"], str(analysis_batches[0].id))

    def test_analysis_batch_can_leave_sample_out_for_analysis(self):
        returned = self.create_and_place_sample("AN-RET", position_id=self.positions[0].id, volume=1.2)
        consumed = self.create_and_place_sample("AN-ZERO", position_id=self.positions[1].id, volume=0.5)

        result = analysis_service.submit_analysis_batch(
            self.session,
            AnalysisBatchCreateInput(
                analysis_type="ELISA",
                performed_at=datetime(2026, 3, 27, 10, 0),
                items=[
                    AnalysisBatchItemInput(sample_id=returned.id, remaining_volume=0.9, thaw_increment=1, returned_to_storage=True),
                    AnalysisBatchItemInput(sample_id=consumed.id, remaining_volume=0.0, thaw_increment=1, returned_to_storage=False),
                ],
            ),
            self.user,
        )

        self.assertEqual(result.returned_count, 1)
        self.assertEqual(result.out_for_analysis_count, 1)

        returned_detail = sample_service.get_sample_detail(self.session, returned.id)
        consumed_detail = sample_service.get_sample_detail(self.session, consumed.id)
        self.assertEqual(returned_detail.custody_label, "in storage")
        self.assertEqual(returned_detail.usage_label, "used")
        self.assertEqual(consumed_detail.custody_label, "out for analysis")
        self.assertEqual(consumed_detail.usage_label, "used")
        self.assertIsNone(consumed_detail.location_label)

    def test_analysis_batch_can_move_return_position(self):
        sample = self.create_and_place_sample("AN-MOVE", position_id=self.positions[0].id, volume=1.0)

        analysis_service.submit_analysis_batch(
            self.session,
            AnalysisBatchCreateInput(
                analysis_type="Sequencing",
                performed_at=datetime(2026, 3, 27, 11, 0),
                items=[
                    AnalysisBatchItemInput(
                        sample_id=sample.id,
                        remaining_volume=0.7,
                        thaw_increment=1,
                        returned_to_storage=True,
                        return_position_id=self.box_two_positions[0].id,
                    )
                ],
            ),
            self.user,
        )

        detail = sample_service.get_sample_detail(self.session, sample.id)
        self.assertIn("Box 2", detail.location_path)
        self.assertEqual(detail.location_label, "A1")
        item = self.session.query(models.AnalysisItem).one()
        self.assertEqual(item.from_position_id, self.positions[0].id)
        self.assertEqual(item.to_position_id, self.box_two_positions[0].id)

    def test_analysis_log_workbook_contains_optional_header_and_prefilled_rows(self):
        sample = self.create_and_place_sample("AN-LOG", position_id=self.positions[0].id, volume=1.2)

        workbook_bytes = analysis_service.generate_analysis_log_xlsx(self.session, [sample.id])
        workbook = load_workbook(filename=BytesIO(workbook_bytes))
        sheet = workbook["Analysis Log"]
        lookup_sheet = workbook["_lists"]

        self.assertEqual(workbook.sheetnames, ["Analysis Log", "_lists"])
        self.assertEqual(sheet["A2"].value, "analysis_type")
        self.assertEqual(sheet["A6"].value, "sample_pk")
        self.assertEqual(sheet["H6"].value, "remaining_volume")
        self.assertEqual(sheet["A7"].value, sample.id)
        self.assertEqual(sheet["B7"].value, "AN-LOG")
        self.assertEqual(sheet["D7"].value, "Box 1")
        self.assertEqual(sheet["E7"].value, "A1")
        self.assertEqual(sheet["F7"].value, 1.2)
        self.assertEqual(sheet["H7"].value, 1.2)
        self.assertEqual(sheet["I7"].value, "yes")
        self.assertEqual(sheet["L7"].value, 1)
        self.assertEqual(lookup_sheet.sheet_state, "hidden")
        self.assertEqual(lookup_sheet["A2"].value, "yes")
        self.assertEqual(lookup_sheet["A3"].value, "no")

    def test_analysis_import_preview_and_commit_from_workbook(self):
        sample = self.create_and_place_sample("AN-XLSX", position_id=self.positions[0].id, volume=1.0)

        workbook = load_workbook(filename=BytesIO(analysis_service.generate_analysis_log_xlsx(self.session, [sample.id])))
        sheet = workbook["Analysis Log"]
        sheet["B2"] = "Route Assay"
        sheet["B3"] = "03/27/26 14:00"
        sheet["B4"] = "Workbook run"
        sheet["H7"] = 0.7
        sheet["L7"] = 2
        buffer = BytesIO()
        workbook.save(buffer)

        raw_payload = analysis_service.analysis_workbook_to_payload(buffer.getvalue())
        preview = analysis_service.preview_analysis_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 1)
        self.assertEqual(preview.invalid_rows, 0)
        self.assertEqual(preview.analysis_type, "Route Assay")
        self.assertEqual(preview.performed_at, "03/27/26 14:00")

        result = analysis_service.commit_analysis_import(
            self.session,
            AnalysisImportCommitInput(raw_payload=raw_payload),
            self.user,
        )
        self.assertEqual(result.imported_rows, 1)
        self.assertEqual(result.batch_id, 1)

        detail = sample_service.get_sample_detail(self.session, sample.id)
        self.assertEqual(detail.custody_label, "in storage")
        self.assertEqual(detail.usage_label, "used")
        self.assertEqual(detail.volume, 0.7)
        self.assertEqual(detail.thaw_count, 2)
        batch = self.session.query(models.AnalysisBatch).one()
        self.assertEqual(batch.analysis_type, "Route Assay")
        self.assertEqual(batch.overall_notes, "Workbook run")

    def test_batch_modify_workbook_contains_current_and_editable_columns(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="MOD-LOG",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="30",
                aliquot_number=2,
                hemolysis_classification=1,
                volume=1.2,
                thaw_count=1,
                notes="Current note",
                collection_at=datetime(2026, 3, 27, 8, 15),
            ),
            self.user,
        )

        workbook_bytes = batch_modify_service.generate_modify_log_xlsx(self.session, [sample.id])
        workbook = load_workbook(filename=BytesIO(workbook_bytes))
        sheet = workbook["Batch Modify"]

        self.assertEqual(sheet["A1"].value, "sample_pk")
        self.assertEqual(sheet["C1"].value, "current_sample_type")
        self.assertEqual(sheet["O1"].value, "sample_type")
        self.assertEqual(sheet["A2"].value, sample.id)
        self.assertEqual(sheet["B2"].value, "MOD-LOG")
        self.assertEqual(sheet["C2"].value, "Plasma")
        self.assertEqual(sheet["D2"].value, "Sovary")
        self.assertEqual(sheet["E2"].value, "current")
        self.assertEqual(sheet["G2"].value, "30")
        self.assertEqual(sheet["O2"].value, "Plasma")
        self.assertEqual(sheet["Q2"].value, "current")
        self.assertEqual(sheet["S2"].value, "30")
        self.assertEqual(sheet["U2"].value, 1)

    def test_batch_modify_preview_and_commit_updates_samples_and_groups_events(self):
        alt_type = admin_service.create_sample_type(
            self.session,
            SampleTypeCreate(name="Serum", description="Serum"),
        )
        alt_study = admin_service.create_study(
            self.session,
            StudyCreate(name="Modify Study"),
        )
        first = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="MOD-GRP-1",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
                volume=1.0,
                thaw_count=0,
                notes="Old note",
                collection_at=datetime(2026, 3, 27, 8, 0),
            ),
            self.user,
        )
        second = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="MOD-GRP-2",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="15",
                aliquot_number=2,
                volume=0.8,
                thaw_count=1,
                notes="Second old note",
                collection_at=datetime(2026, 3, 27, 8, 30),
            ),
            self.user,
        )

        workbook = load_workbook(filename=BytesIO(batch_modify_service.generate_modify_log_xlsx(self.session, [first.id, second.id])))
        sheet = workbook["Batch Modify"]
        sheet["O2"] = "Serum"
        sheet["P2"] = "Modify Study"
        sheet["Q2"] = "current"
        sheet["R2"] = "2"
        sheet["S2"] = "60"
        sheet["T2"] = 3
        sheet["U2"] = 4
        sheet["V2"] = 0.7
        sheet["W2"] = 2
        sheet["X2"] = "Updated note"
        sheet["Y2"] = "03/28/26 10:15"
        sheet["X3"] = "Second updated note"
        sheet["W3"] = 5
        buffer = BytesIO()
        workbook.save(buffer)

        raw_payload = batch_modify_service.modify_workbook_to_payload(buffer.getvalue())
        preview = batch_modify_service.preview_modify_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 2)
        self.assertEqual(preview.invalid_rows, 0)

        result = batch_modify_service.commit_modify_import(
            self.session,
            BatchModifyImportCommitInput(raw_payload=raw_payload),
            self.user,
        )
        self.assertEqual(result.imported_rows, 2)
        self.assertIsNotNone(result.batch_group_id)

        first_detail = sample_service.get_sample_detail(self.session, first.id)
        second_detail = sample_service.get_sample_detail(self.session, second.id)
        self.assertEqual(first_detail.sample_type_name, "Serum")
        self.assertEqual(first_detail.study_name, "Modify Study")
        self.assertEqual(first_detail.visit_label, "2")
        self.assertEqual(first_detail.timepoint_label, "60")
        self.assertEqual(first_detail.aliquot_number, 3)
        self.assertEqual(first_detail.hemolysis_classification, 4)
        self.assertEqual(first_detail.volume, 0.7)
        self.assertEqual(first_detail.thaw_count, 2)
        self.assertEqual(first_detail.notes, "Updated note")
        self.assertEqual(second_detail.thaw_count, 5)
        self.assertEqual(second_detail.notes, "Second updated note")

        feed = event_service.list_events(self.session, query=EventSearchQuery(event_type="update_sample", sample_query="MOD-GRP", limit=20))
        self.assertEqual(len(feed), 3)
        self.assertTrue(feed[0].is_group_parent)
        self.assertEqual(feed[0].group_kind, "batch_modify")
        self.assertEqual(feed[0].action_label, "Batch Modify")
        self.assertEqual(feed[0].group_count, 2)
        self.assertTrue(all(item.is_group_child for item in feed[1:]))

    def test_batch_modify_preview_rejects_invalid_metadata_and_identity_collisions(self):
        colliding = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="MOD-DUP",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="60",
                aliquot_number=1,
                thaw_count=0,
            ),
            self.user,
        )
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="MOD-DUP",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
                thaw_count=0,
            ),
            self.user,
        )
        self.assertIsNotNone(colliding)

        workbook = load_workbook(filename=BytesIO(batch_modify_service.generate_modify_log_xlsx(self.session, [sample.id])))
        sheet = workbook["Batch Modify"]
        sheet["O2"] = "Missing Type"
        sheet["Q2"] = "bad_role"
        sheet["R2"] = "V1"
        sheet["S2"] = "60"
        sheet["U2"] = 9
        sheet["W2"] = ""
        buffer = BytesIO()
        workbook.save(buffer)

        raw_payload = batch_modify_service.modify_workbook_to_payload(buffer.getvalue())
        preview = batch_modify_service.preview_modify_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 0)
        self.assertEqual(preview.invalid_rows, 1)
        self.assertIn("Sample type was not found", preview.rows[0].errors)
        self.assertIn("Study role must be current or retired", preview.rows[0].errors)
        self.assertIn("Visit must contain numbers only", preview.rows[0].errors)
        self.assertIn("Hemolysis must be a whole number from 0 to 6", preview.rows[0].errors)
        self.assertIn("Thaw count is required", preview.rows[0].errors)

        sheet["O2"] = "Plasma"
        sheet["Q2"] = "current"
        sheet["R2"] = "1"
        sheet["U2"] = 2
        sheet["W2"] = 1
        buffer = BytesIO()
        workbook.save(buffer)
        collision_payload = batch_modify_service.modify_workbook_to_payload(buffer.getvalue())
        collision_preview = batch_modify_service.preview_modify_import(self.session, collision_payload)
        self.assertEqual(collision_preview.valid_rows, 0)
        self.assertIn("This ID, type, visit, timepoint, and aliquot combination already exists", collision_preview.rows[0].errors)

        result = batch_modify_service.commit_modify_import(
            self.session,
            BatchModifyImportCommitInput(raw_payload=collision_payload),
            self.user,
        )
        self.assertEqual(result.imported_rows, 0)
        self.assertIsNone(result.batch_group_id)
        self.assertTrue(result.global_errors)

    def test_analysis_import_preview_rejects_archived_duplicate_and_malformed_rows(self):
        sample = self.create_and_place_sample("AN-BAD", position_id=self.positions[0].id, volume=1.0)
        workbook = load_workbook(filename=BytesIO(analysis_service.generate_analysis_log_xlsx(self.session, [sample.id])))
        sheet = workbook["Analysis Log"]
        sheet["I7"] = "maybe"
        for column in range(1, len(analysis_service.HEADERS) + 1):
            sheet.cell(row=8, column=column).value = sheet.cell(row=7, column=column).value
        buffer = BytesIO()
        workbook.save(buffer)
        raw_payload = analysis_service.analysis_workbook_to_payload(buffer.getvalue())

        sample_service.retrieve_sample(self.session, sample.id, RetrieveSampleInput(note="Archived after log"), self.user)

        preview = analysis_service.preview_analysis_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 0)
        self.assertEqual(preview.invalid_rows, 2)
        self.assertIn("Sample must be placed and not archived before analysis", preview.rows[0].errors)
        self.assertIn("sample_pk is duplicated in this workbook", preview.rows[0].errors)
        self.assertIn("returned_to_storage must be yes or no", preview.rows[0].errors)

        result = analysis_service.commit_analysis_import(
            self.session,
            AnalysisImportCommitInput(raw_payload=raw_payload),
            self.user,
        )
        self.assertEqual(result.imported_rows, 0)
        self.assertIsNone(result.batch_id)
        self.assertTrue(result.global_errors)

    def test_analysis_batch_is_atomic_on_invalid_selection(self):
        valid = self.create_and_place_sample("AN-VALID", position_id=self.positions[0].id, volume=1.0)
        invalid = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="AN-UNPLACED", sample_type_id=self.sample_type.id, study_id=self.study.id, volume=0.6),
            self.user,
        )

        with self.assertRaises(analysis_service.AnalysisError):
            analysis_service.submit_analysis_batch(
                self.session,
                AnalysisBatchCreateInput(
                    analysis_type="Mass Spec",
                    performed_at=datetime(2026, 3, 27, 12, 0),
                    items=[
                        AnalysisBatchItemInput(sample_id=valid.id, remaining_volume=0.8, thaw_increment=1, returned_to_storage=True),
                        AnalysisBatchItemInput(sample_id=invalid.id, remaining_volume=0.4, thaw_increment=1, returned_to_storage=True),
                    ],
                ),
                self.user,
            )

        valid_detail = sample_service.get_sample_detail(self.session, valid.id)
        invalid_detail = sample_service.get_sample_detail(self.session, invalid.id)
        self.assertEqual(valid_detail.study_role, "current")
        self.assertEqual(valid_detail.custody_label, "in storage")
        self.assertEqual(valid_detail.usage_label, "unused")
        self.assertEqual(valid_detail.volume, 1.0)
        self.assertEqual(valid_detail.location_label, "A1")
        self.assertEqual(invalid_detail.study_role, "current")
        self.assertEqual(invalid_detail.custody_label, "unplaced")
        self.assertEqual(self.session.query(models.AnalysisBatch).count(), 0)
        self.assertEqual(self.session.query(models.AnalysisItem).count(), 0)

    def test_analysis_preview_marks_unplaced_and_archived_samples_ineligible(self):
        placed = self.create_and_place_sample("AN-PREVIEW", position_id=self.positions[0].id, volume=1.0)
        archived = self.create_and_place_sample("AN-ARCH", position_id=self.positions[1].id, volume=0.5)
        sample_service.retrieve_sample(
            self.session,
            archived.id,
            RetrieveSampleInput(note="Consumed"),
            self.user,
        )
        unplaced = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="AN-UNPLACED-2", sample_type_id=self.sample_type.id, study_id=self.study.id, volume=0.7),
            self.user,
        )

        preview = analysis_service.preview_samples(
            self.session,
            AnalysisPreviewRequest(sample_ids=[placed.id, archived.id, unplaced.id, 9999]),
        )

        self.assertEqual(len(preview.samples), 4)
        self.assertTrue(preview.samples[0].eligible)
        self.assertFalse(preview.samples[1].eligible)
        self.assertIn("Archived", preview.samples[1].ineligibility_reason)
        self.assertFalse(preview.samples[2].eligible)
        self.assertIn("placed in storage", preview.samples[2].ineligibility_reason)
        self.assertFalse(preview.samples[3].eligible)
        self.assertIn("not found", preview.samples[3].ineligibility_reason)

    def test_global_feed_groups_analysis_batch_but_sample_history_stays_flat(self):
        first = self.create_and_place_sample("AN-GRP-1", position_id=self.positions[0].id, volume=1.0)
        second = self.create_and_place_sample("AN-GRP-2", position_id=self.positions[1].id, volume=1.0)

        analysis_service.submit_analysis_batch(
            self.session,
            AnalysisBatchCreateInput(
                analysis_type="Grouped Assay",
                performed_at=datetime(2026, 3, 27, 15, 0),
                overall_notes="Grouped batch",
                items=[
                    AnalysisBatchItemInput(sample_id=first.id, remaining_volume=0.8, thaw_increment=1, returned_to_storage=True),
                    AnalysisBatchItemInput(sample_id=second.id, remaining_volume=0.0, thaw_increment=1, returned_to_storage=False),
                ],
            ),
            self.user,
        )

        feed = event_service.list_events(self.session, query=EventSearchQuery(event_type="analyze_sample", limit=20))
        self.assertEqual(len(feed), 3)
        parent = feed[0]
        children = feed[1:]
        self.assertTrue(parent.is_group_parent)
        self.assertEqual(parent.group_kind, "analysis_batch")
        self.assertEqual(parent.group_count, 2)
        self.assertEqual(parent.title, "Grouped Assay batch #1")
        self.assertIn("2 samples", parent.context_line)
        self.assertEqual(len(parent.group_children), 2)
        self.assertTrue(any(item.sample_identifier == "AN-GRP-1" and item.disposition == "Returned" for item in parent.group_children))
        self.assertTrue(any(item.sample_identifier == "AN-GRP-2" and item.disposition == "Out for analysis" for item in parent.group_children))
        self.assertTrue(all(child.is_group_child for child in children))
        self.assertEqual({child.sample_identifier for child in children}, {"AN-GRP-1", "AN-GRP-2"})

        first_detail = sample_service.get_sample_detail(self.session, first.id)
        first_analysis_events = [event for event in first_detail.events if event.event_type == "analyze_sample"]
        self.assertEqual(len(first_analysis_events), 1)
        self.assertFalse(first_analysis_events[0].is_group_parent)
        self.assertFalse(first_analysis_events[0].is_group_child)

    def test_feed_groups_generic_batch_events_with_explicit_group_metadata(self):
        first = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="GEN-GRP-1", sample_type_id=self.sample_type.id, study_id=self.study.id),
            self.user,
        )
        second = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="GEN-GRP-2", sample_type_id=self.sample_type.id, study_id=self.study.id),
            self.user,
        )
        for sample in (first, second):
            event = models.Event(
                event_type=models.EventType.update_sample,
                user_id=self.user.id,
                sample_id=sample.id,
                created_at=datetime.utcnow(),
            )
            event.set_payload(
                {
                    "batch_group_kind": "bulk_adjustment",
                    "batch_group_id": "demo-1",
                    "batch_group_title": "Bulk adjustment demo",
                    "changes": [{"field": "notes", "label": "Notes", "before": "--", "after": "Adjusted"}],
                    "after": {"sample_id": sample.sample_id, "type": "Plasma", "study": "Sovary", "visit": "--", "timepoint": "--", "aliquot": "--", "hemolysis": "--", "study_role": "current", "custody": "unplaced", "usage": "unused", "volume": "--", "thaw_count": "0", "notes": "Adjusted", "collection": "--"},
                }
            )
            self.session.add(event)
        self.session.commit()

        feed = event_service.list_events(self.session, query=EventSearchQuery(sample_query="GEN-GRP", limit=20))
        grouped = [event for event in feed if event.is_group_parent]
        children = [event for event in feed if event.is_group_child]
        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0].group_kind, "bulk_adjustment")
        self.assertEqual(grouped[0].group_id, "demo-1")
        self.assertEqual(grouped[0].group_count, 2)
        self.assertEqual(len(grouped[0].group_children), 2)
        self.assertEqual(len(children), 2)
        self.assertEqual({child.sample_identifier for child in children}, {"GEN-GRP-1", "GEN-GRP-2"})

    def test_analysis_batch_respects_manual_thaw_increment(self):
        sample = self.create_and_place_sample("AN-THAW", position_id=self.positions[0].id, volume=1.0)

        analysis_service.submit_analysis_batch(
            self.session,
            AnalysisBatchCreateInput(
                analysis_type="Proteomics",
                performed_at=datetime(2026, 3, 27, 13, 0),
                items=[
                    AnalysisBatchItemInput(sample_id=sample.id, remaining_volume=0.75, thaw_increment=3, returned_to_storage=True)
                ],
            ),
            self.user,
        )

        detail = sample_service.get_sample_detail(self.session, sample.id)
        self.assertEqual(detail.thaw_count, 3)

    def test_admin_can_delete_any_sample_and_preserve_delete_audit(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-DELETE",
                study_id=self.study.id,
                sample_type_id=self.sample_type.id,
                thaw_count=1,
            ),
            self.user,
        )
        sample_service.place_sample(
            self.session,
            sample.id,
            PlaceSampleInput(position_id=self.positions[0].id),
            self.user,
        )

        sample_service.delete_sample(self.session, sample.id, self.user)

        self.assertIsNone(sample_service.get_sample_detail(self.session, sample.id))
        refreshed_position = storage_service.get_position(self.session, self.positions[0].id)
        self.assertIsNone(refreshed_position.location)
        events = event_service.list_events(self.session, limit=20)
        delete_event = next(event for event in events if event.event_type == "delete_sample")
        self.assertEqual(delete_event.sample_id, None)
        self.assertEqual(delete_event.sample_identifier, "S-DELETE")
        self.assertEqual(delete_event.payload["location_label"], "A1")
        self.assertEqual(delete_event.title, "S-DELETE Plasma")

    def test_event_feed_formats_rows_and_supports_filters(self):
        operator = auth_service.sync_user(self.session, "operator", "Operator")
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-ACT-1",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
            ),
            operator,
        )
        sample_service.add_note_entry(
            self.session,
            sample.id,
            SampleNoteCreateInput(text="Should stay off the global feed"),
            operator,
        )
        sample_service.place_sample(
            self.session,
            sample.id,
            PlaceSampleInput(position_id=self.positions[0].id),
            operator,
        )
        sample_service.retrieve_sample(
            self.session,
            sample.id,
            RetrieveSampleInput(note="Archived after review"),
            operator,
        )

        feed = event_service.list_events(
            self.session,
            query=EventSearchQuery(
                sample_query="S-ACT-1",
                user_id=operator.id,
                limit=20,
            ),
        )

        self.assertTrue(feed)
        self.assertNotIn("add_note", {event.event_type for event in feed})
        archive_event = next(event for event in feed if event.event_type == "status_change")
        self.assertEqual(archive_event.risk_level, "high")
        self.assertTrue(archive_event.is_high_risk)
        self.assertEqual(archive_event.related_url, f"/samples/{sample.id}")
        self.assertEqual(archive_event.display_action, "update_sample")
        self.assertEqual(archive_event.title, "S-ACT-1 Plasma V1 T00 #1")
        self.assertTrue(any(item.label == "Changed" for item in archive_event.pill_items))
        self.assertIsNone(archive_event.context_line)

        placement_events = event_service.list_events(
            self.session,
            query=EventSearchQuery(event_type="place_sample", sample_query="S-ACT-1", limit=20),
        )
        self.assertEqual(len(placement_events), 1)
        self.assertEqual(placement_events[0].title, "S-ACT-1 Plasma V1 T00 #1")
        self.assertIn("Box 1", placement_events[0].context_line)
        self.assertIsNone(placement_events[0].drawer_context_line)
        self.assertEqual(placement_events[0].detail_sections[0].title, "Location")
        self.assertEqual(placement_events[0].detail_sections[0].items[0].label, "")
        self.assertEqual(placement_events[0].detail_sections[0].layout, "single_value")

    def test_update_event_records_before_and_after_metadata(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-AUDIT",
                sample_type_id=self.sample_type.id,
                study_id=self.study.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
                volume=1.0,
                notes="Initial note",
            ),
            self.user,
        )

        second_study = admin_service.create_study(self.session, StudyCreate(name="Follow Up"))
        sample_service.update_sample(
            self.session,
            sample.id,
            SampleUpdateInput(
                study_id=second_study.id,
                timepoint_label="30",
                volume=0.8,
                notes="Updated note",
            ),
            self.user,
        )

        update_event = next(
            event
            for event in event_service.list_events(self.session, query=EventSearchQuery(sample_query="S-AUDIT", limit=20))
            if event.event_type == "update_sample"
        )

        self.assertEqual(update_event.title, "S-AUDIT Plasma V1 T30 #1")
        self.assertTrue(any(change.label == "Study" and change.before == "Sovary" and change.after == "Follow Up" for change in update_event.change_items))
        self.assertTrue(any(change.label == "Timepoint" and change.before == "00" and change.after == "30" for change in update_event.change_items))
        self.assertTrue(any(change.label == "Volume" and change.before == "1 mL" and change.after == "0.8 mL" for change in update_event.change_items))
        self.assertIn("study", update_event.raw_payload["before"])
        self.assertIn("study", update_event.raw_payload["after"])
        self.assertFalse(update_event.has_legacy_detail_gap)
        self.assertEqual(update_event.metadata_section.title, "Current metadata")

    def test_legacy_update_event_flags_missing_detailed_tracking(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="S-LEGACY"),
            self.user,
        )
        legacy_event = models.Event(
            event_type=models.EventType.update_sample,
            user_id=self.user.id,
            sample_id=sample.id,
            created_at=datetime.utcnow(),
        )
        legacy_event.set_payload({"updated_at": datetime.utcnow().isoformat()})
        self.session.add(legacy_event)
        self.session.commit()

        event = next(
            event
            for event in event_service.list_events(self.session, query=EventSearchQuery(sample_query="S-LEGACY", limit=20))
            if event.id == legacy_event.id
        )

        self.assertTrue(event.has_legacy_detail_gap)
        self.assertEqual(event.title, "S-LEGACY")

    def test_composite_audit_title_omits_blank_metadata(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-PARTIAL",
                sample_type_id=self.sample_type.id,
            ),
            self.user,
        )

        create_event = next(
            event
            for event in event_service.list_events(self.session, query=EventSearchQuery(sample_query="S-PARTIAL", limit=20))
            if event.event_type == "create_sample"
        )

        self.assertEqual(create_event.title, "S-PARTIAL Plasma")
        self.assertIsNone(create_event.context_line)

    def test_composite_audit_title_prefixes_numeric_visit_and_timepoint(self):
        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-NUMERIC",
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="120",
                aliquot_number=1,
            ),
            self.user,
        )

        create_event = next(
            event
            for event in event_service.list_events(self.session, query=EventSearchQuery(sample_query="S-NUMERIC", limit=20))
            if event.event_type == "create_sample"
        )

        self.assertEqual(create_event.title, "S-NUMERIC Plasma V1 T120 #1")

    def test_box_creation_emits_single_create_box_activity(self):
        new_box = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Box 55", node_type="box", parent_id=self.rack.id),
            self.user,
        )
        storage_service.create_box_positions(
            self.session,
            BoxCreateInput(box_id=new_box.id, rows=3, cols=4),
            self.user,
        )

        events = event_service.list_events(self.session, query=EventSearchQuery(limit=50))
        create_box_events = [event for event in events if event.display_action == "create_box" and event.title == "Box 55"]
        self.assertEqual(len(create_box_events), 1)
        self.assertIn("3 x 4", create_box_events[0].context_line)
        self.assertIn("12 positions", create_box_events[0].context_line)
        self.assertIsNone(create_box_events[0].drawer_context_line)
        self.assertEqual(create_box_events[0].detail_sections[0].layout, "compact")

        filtered = event_service.list_events(
            self.session,
            query=EventSearchQuery(event_type="create_box", limit=20),
        )
        self.assertTrue(any(event.title == "Box 55" for event in filtered))
        self.assertTrue(all(event.display_action == "create_box" for event in filtered))

    def test_prevent_duplicate_occupancy(self):
        first = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="S-100", study_id=self.study.id),
            self.user,
        )
        second = sample_service.create_sample(
            self.session,
            SampleCreateInput(sample_id="S-101", study_id=self.study.id),
            self.user,
        )
        sample_service.place_sample(
            self.session,
            first.id,
            PlaceSampleInput(position_id=self.positions[0].id),
            self.user,
        )

        with self.assertRaises(sample_service.SampleError):
            sample_service.place_sample(
                self.session,
                second.id,
                PlaceSampleInput(position_id=self.positions[0].id),
                self.user,
            )

    def test_search_filters_distinguish_visit_and_timepoint(self):
        first = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="SOV21-V1-T60",
                study_id=self.study.id,
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="60",
                thaw_count=1,
                hemolysis_classification=4,
                volume=0.8,
                notes="First visit",
            ),
            self.user,
        )
        second = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="SOV21-V2-T60",
                study_id=self.study.id,
                sample_type_id=self.sample_type.id,
                visit_label="2",
                timepoint_label="60",
                volume=1.0,
            ),
            self.user,
        )
        sample_service.place_sample(
            self.session,
            first.id,
            PlaceSampleInput(position_id=self.positions[0].id),
            self.user,
        )

        visit_results = sample_service.search_samples(
            self.session,
            SampleSearchQuery(study_id=self.study.id, visit_label="1", timepoint_label="60"),
        )
        self.assertEqual(len(visit_results), 1)
        self.assertEqual(visit_results[0].sample_id, "SOV21-V1-T60")

        text_results = sample_service.search_samples(self.session, SampleSearchQuery(q="T60"))
        self.assertEqual(len(text_results), 2)

        type_results = sample_service.search_samples(self.session, SampleSearchQuery(q="Plasma"))
        self.assertEqual(len(type_results), 2)

        location_results = sample_service.search_samples(self.session, SampleSearchQuery(q="Box 1"))
        self.assertEqual(len(location_results), 1)
        self.assertEqual(location_results[0].sample_id, "SOV21-V1-T60")

        freezer_results = sample_service.search_samples(
            self.session,
            SampleSearchQuery(storage_node_ids=[self.freezer.id]),
        )
        self.assertEqual(len(freezer_results), 1)
        self.assertEqual(freezer_results[0].sample_id, "SOV21-V1-T60")

        box_results = sample_service.search_samples(
            self.session,
            SampleSearchQuery(storage_node_ids=[self.box.id]),
        )
        self.assertEqual(len(box_results), 1)
        self.assertEqual(box_results[0].sample_id, "SOV21-V1-T60")

        filtered = sample_service.search_samples(
            self.session,
            SampleSearchQuery(usage="used", thaw_count_min=1, volume_max=0.9, hemolysis_min=4),
        )
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0].sample_id, "SOV21-V1-T60")
        self.assertEqual(filtered[0].hemolysis_classification, 4)

        option_payload = sample_service.get_filter_options(
            self.session,
            SampleSearchQuery(q="SOV21", usages=["used", "unused"]),
            "usage",
        )
        self.assertEqual({option.value: option.count for option in option_payload.options}, {"unused": 1, "used": 1})

    def test_hemolysis_must_be_zero_through_six(self):
        with self.assertRaises(sample_service.SampleError):
            sample_service.create_sample(
                self.session,
                SampleCreateInput(
                    sample_id="S-BAD-HEM",
                    sample_type_id=self.sample_type.id,
                    hemolysis_classification=7,
                ),
                self.user,
            )

    def test_visit_and_timepoint_must_be_numeric_only(self):
        with self.assertRaises(sample_service.SampleError):
            sample_service.create_sample(
                self.session,
                SampleCreateInput(
                    sample_id="S-BAD-VISIT",
                    sample_type_id=self.sample_type.id,
                    visit_label="V1",
                ),
                self.user,
            )

        sample = sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-GOOD-VISIT",
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="60",
            ),
            self.user,
        )

        with self.assertRaises(sample_service.SampleError):
            sample_service.update_sample(
                self.session,
                sample.id,
                SampleUpdateInput(timepoint_label="T60"),
                self.user,
            )

    def test_storage_hierarchy_and_nickname_rules(self):
        with self.assertRaises(storage_service.StorageError):
            storage_service.create_storage_node(
                self.session,
                StorageNodeCreate(name="Shelf root", node_type="shelf"),
                self.user,
            )

        rack = storage_service.update_storage_node(
            self.session,
            self.rack.id,
            StorageNodeUpdate(name="Rack 1", nickname="Cold lane", notes="North corner"),
            self.user,
        )
        self.assertEqual(rack.nickname, "Cold lane")
        self.assertEqual(rack.notes, "North corner")

        moved = storage_service.move_storage_node(
            self.session,
            self.rack.id,
            StorageNodeMoveInput(parent_id=self.freezer.id),
            self.user,
        )
        self.assertEqual(moved.parent_id, self.freezer.id)

        with self.assertRaises(storage_service.StorageError):
            storage_service.move_storage_node(
                self.session,
                self.box.id,
                StorageNodeMoveInput(parent_id=self.freezer.id),
                self.user,
            )

    def test_unique_names_for_freezers_and_boxes(self):
        with self.assertRaises(storage_service.StorageError):
            storage_service.create_storage_node(
                self.session,
                StorageNodeCreate(name="Freezer A", node_type="freezer"),
                self.user,
            )

        other_rack = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Rack 2", node_type="rack", parent_id=self.shelf.id),
            self.user,
        )
        with self.assertRaises(storage_service.StorageError):
            storage_service.create_storage_node(
                self.session,
                StorageNodeCreate(name="Box 1", node_type="box", parent_id=other_rack.id),
                self.user,
            )

    def test_bulk_sample_import_preview_flags_duplicates_and_invalid_values(self):
        sample_service.create_sample(
            self.session,
            SampleCreateInput(
                sample_id="S-001",
                study_id=self.study.id,
                sample_type_id=self.sample_type.id,
                visit_label="1",
                timepoint_label="00",
                aliquot_number=1,
            ),
            self.user,
        )
        raw_payload = "\n".join(
            [
                "sample_id,sample_type,study,visit,timepoint,aliquot,hemolysis,study_role,volume,volume_units,thaw_count,notes,collection_at,position",
                "S-001,Plasma,Sovary,1,00,1,1,current,1.0,mL,0,Existing combination,03/10/26 08:00,A1",
                "S-002,Unknown,Sovary,V1,T15,alpha,8,bad_role,-1,mL,-1,Invalid row,not-a-date,A2",
                "S-003,Plasma,MISSING,1,30,2,2,current,0.5,mL,1,Missing study,03/10/26 08:30,A3",
                "S-004,Plasma,Sovary,2,15,1,2,current,0.5,mL,1,Duplicate composite one,03/10/26 08:45,A4",
                "S-004,Plasma,Sovary,2,15,1,3,current,0.4,mL,1,Duplicate composite two,03/10/26 09:00,B1",
            ]
        )

        preview = bulk_import_service.preview_sample_import(self.session, raw_payload, self.box.id)

        self.assertEqual(preview.total_rows, 5)
        self.assertEqual(preview.valid_rows, 0)
        self.assertEqual(preview.invalid_rows, 5)
        self.assertIn("This ID, type, visit, timepoint, and aliquot combination already exists", preview.rows[0].errors)
        self.assertIn("Sample type was not found", preview.rows[1].errors)
        self.assertIn("Visit must contain numbers only", preview.rows[1].errors)
        self.assertIn("Timepoint must contain numbers only", preview.rows[1].errors)
        self.assertIn("Hemolysis must be a whole number from 0 to 6", preview.rows[1].errors)
        self.assertIn("Study role must be current or retired", preview.rows[1].errors)
        self.assertIn("Study was not found", preview.rows[2].errors)
        self.assertIn("This ID, type, visit, timepoint, and aliquot combination is duplicated in this import", preview.rows[3].errors)
        self.assertIn("This ID, type, visit, timepoint, and aliquot combination is duplicated in this import", preview.rows[4].errors)

    def test_bulk_sample_import_requires_sample_type_and_collection_date(self):
        raw_payload = "\n".join(
            [
                "sample_id,sample_type,study,visit,timepoint,aliquot,study_role,volume,volume_units,thaw_count,notes,collection_at,box,position",
                "S-010,,Sovary,1,00,1,current,1.0,mL,0,Missing requireds,,Box 1,A1",
            ]
        )

        preview = bulk_import_service.preview_sample_import(self.session, raw_payload)

        self.assertEqual(preview.valid_rows, 0)
        self.assertIn("Sample type is required", preview.rows[0].errors)
        self.assertIn("Collection date is required", preview.rows[0].errors)

    def test_bulk_sample_import_commit_places_explicit_and_sequential_positions(self):
        raw_payload = "\n".join(
            [
                "sample_id,sample_type,study,visit,timepoint,aliquot,study_role,volume,volume_units,thaw_count,notes,collection_at,position",
                "S-200,Plasma,Sovary,1,00,1,current,1.0,mL,0,Explicit placement,03/10/26 08:00,A2",
                "S-201,Plasma,Sovary,1,15,2,current,0.8,mL,0,Sequential placement,03/10/26 08:15,",
            ]
        )

        preview = bulk_import_service.preview_sample_import(self.session, raw_payload, self.box.id)
        self.assertEqual(preview.valid_rows, 2)
        self.assertEqual(preview.rows[0].assigned_position, "A2")
        self.assertEqual(preview.rows[1].assigned_position, "A1")

        result = bulk_import_service.commit_sample_import(
            self.session,
            BulkSampleImportCommitInput(raw_payload=raw_payload, target_box_id=self.box.id),
            self.user,
        )

        self.assertEqual(result.imported_rows, 2)
        explicit = sample_service.get_sample_detail(
            self.session,
            sample_service.search_samples(self.session, SampleSearchQuery(q="S-200"))[0].id,
        )
        sequential = sample_service.get_sample_detail(
            self.session,
            sample_service.search_samples(self.session, SampleSearchQuery(q="S-201"))[0].id,
        )
        self.assertEqual(explicit.location_label, "A2")
        self.assertEqual(sequential.location_label, "A1")

    def test_bulk_sample_import_uses_row_box_name_with_page_level_fallback(self):
        raw_payload = "\n".join(
            [
                "sample_id,sample_type,study,visit,timepoint,aliquot,study_role,volume,volume_units,thaw_count,notes,collection_at,box,position",
                "S-300,Plasma,Sovary,1,00,1,current,1.0,mL,0,Row box explicit,03/10/26 08:00,Box 2,A2",
                "S-301,Plasma,Sovary,1,15,2,current,0.8,mL,0,Row box sequential,03/10/26 08:15,Box 2,",
                "S-302,Plasma,Sovary,1,30,3,current,0.7,mL,0,Fallback box,03/10/26 08:30,,",
            ]
        )

        preview = bulk_import_service.preview_sample_import(self.session, raw_payload, self.box.id)
        self.assertEqual(preview.valid_rows, 3)
        self.assertEqual(preview.rows[0].assigned_box_name, "Box 2")
        self.assertEqual(preview.rows[0].assigned_position, "A2")
        self.assertEqual(preview.rows[1].assigned_box_name, "Box 2")
        self.assertEqual(preview.rows[1].assigned_position, "A1")
        self.assertEqual(preview.rows[2].assigned_box_name, "Box 1")
        self.assertEqual(preview.rows[2].assigned_position, "A1")

        result = bulk_import_service.commit_sample_import(
            self.session,
            BulkSampleImportCommitInput(raw_payload=raw_payload, target_box_id=self.box.id),
            self.user,
        )

        self.assertEqual(result.imported_rows, 3)
        row_box_explicit = sample_service.get_sample_detail(
            self.session,
            sample_service.search_samples(self.session, SampleSearchQuery(q="S-300"))[0].id,
        )
        row_box_sequential = sample_service.get_sample_detail(
            self.session,
            sample_service.search_samples(self.session, SampleSearchQuery(q="S-301"))[0].id,
        )
        fallback = sample_service.get_sample_detail(
            self.session,
            sample_service.search_samples(self.session, SampleSearchQuery(q="S-302"))[0].id,
        )
        self.assertIn("Box 2", row_box_explicit.location_path)
        self.assertEqual(row_box_explicit.location_label, "A2")
        self.assertIn("Box 2", row_box_sequential.location_path)
        self.assertEqual(row_box_sequential.location_label, "A1")
        self.assertIn("Box 1", fallback.location_path)
        self.assertEqual(fallback.location_label, "A1")

    def test_sample_excel_template_contains_header_comments_and_roundtrips(self):
        workbook_bytes = bulk_import_service.sample_template_xlsx(
            sample_types=["Plasma", "Serum"],
            studies=["Sovary", "NRS"],
            boxes=["Box 1", "Box 2"],
        )
        workbook = load_workbook(filename=BytesIO(workbook_bytes))
        sheet = workbook["Sample Import"]
        lookup_sheet = workbook["_lists"]

        self.assertEqual(sheet["A1"].value, "sample_id")
        self.assertEqual(workbook.sheetnames, ["Sample Import", "_lists"])
        self.assertEqual(lookup_sheet.sheet_state, "hidden")
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertEqual(sheet["A1"].alignment.vertical, "center")
        self.assertAlmostEqual(sheet.column_dimensions["A"].width, 11.265625)
        self.assertAlmostEqual(sheet.column_dimensions["G"].width, 10.5)
        self.assertAlmostEqual(sheet.column_dimensions["M"].width, 14.0)
        self.assertIsNotNone(sheet["A1"].comment)
        self.assertIn("participant ID", sheet["A1"].comment.text)
        self.assertEqual(sheet["G1"].value, "hemolysis")
        self.assertIn("0 to 6 scale", sheet["G1"].comment.text)
        self.assertEqual(sheet["A2"].alignment.horizontal, "center")
        self.assertEqual(sheet["A2"].alignment.vertical, "center")
        self.assertIsNone(sheet["A2"].comment)
        self.assertEqual(sheet["L2"].alignment.horizontal, "left")
        self.assertEqual(sheet["L2"].alignment.vertical, "top")
        self.assertIsNone(sheet["L2"].comment)
        self.assertEqual(sheet["N1"].value, "box")
        self.assertIsNotNone(sheet["N1"].comment)
        self.assertIn("Exact box name", sheet["N1"].comment.text)
        self.assertIsNone(sheet["A2"].value)
        self.assertIsNone(sheet["M2"].value)
        validations = list(sheet.data_validations.dataValidation)
        validation_formulas = {validation.formula1 for validation in validations}
        self.assertIn("'_lists'!$A$2:$A$3", validation_formulas)
        self.assertIn("'_lists'!$B$2:$B$3", validation_formulas)
        self.assertIn("'_lists'!$C$2:$C$3", validation_formulas)
        self.assertEqual(lookup_sheet["C2"].value, "Box 1")
        self.assertEqual(lookup_sheet["C3"].value, "Box 2")
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            vml = archive.read("xl/drawings/commentsDrawing1.vml").decode("utf-8")
        self.assertIn("width:190px;height:70px", vml)
        self.assertIn("width:560px;height:110px", vml)
        self.assertIn("width:540px;height:85px", vml)

        sheet["A3"] = "S-900"
        sheet["B3"] = "Plasma"
        sheet["C3"] = "Sovary"
        sheet["G3"] = 3
        sheet["M3"] = "03/15/26 14:30"
        sheet["N3"] = "Box 1"
        sheet["O3"] = "A2"
        buffer = BytesIO()
        workbook.save(buffer)

        raw_payload = bulk_import_service.sample_workbook_to_csv(buffer.getvalue())
        self.assertIn("sample_id,sample_type,study,visit,timepoint,aliquot,hemolysis", raw_payload)
        self.assertIn("S-900,Plasma,Sovary,,,,3", raw_payload)

    def test_bulk_box_import_creates_box_under_existing_shelf(self):
        raw_payload = "\n".join(
            [
                "parent,box,rows,cols,box_nickname,notes",
                "Freezer A > Shelf 1,Box 20,2,3,Study set,Created in bulk",
            ]
        )

        preview = bulk_import_service.preview_box_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 1)
        self.assertEqual(preview.rows[0].parent, "Freezer A > Shelf 1")

        result = bulk_import_service.commit_box_import(
            self.session,
            BulkBoxImportCommitInput(raw_payload=raw_payload),
            self.user,
        )

        self.assertEqual(result.imported_rows, 1)
        nodes = storage_service.list_all_nodes(self.session)
        box = next(node for node in nodes if node.node_type.value == "box" and node.name == "Box 20")
        self.assertEqual(box.parent_id, self.shelf.id)
        self.assertEqual(box.nickname, "Study set")
        self.assertEqual(box.notes, "Created in bulk")
        self.assertEqual(len(storage_service.get_box_view(self.session, box.id).positions), 6)

    def test_bulk_box_import_accepts_shelf_and_rack_parent_paths(self):
        existing_rack = storage_service.create_storage_node(
            self.session,
            StorageNodeCreate(name="Rack Path", node_type="rack", parent_id=self.shelf.id),
            self.user,
        )
        self.assertIsNotNone(existing_rack)
        raw_payload = "\n".join(
            [
                "parent,box,rows,cols,box_nickname,notes",
                "Freezer A > Shelf 1,Box Path Shelf,2,2,,Created from shelf path",
                "Freezer A > Shelf 1 > Rack Path,Box Path Rack,2,2,,Created from rack path",
            ]
        )

        preview = bulk_import_service.preview_box_import(self.session, raw_payload)

        self.assertEqual(preview.valid_rows, 2)
        self.assertEqual(preview.rows[0].parent, "Freezer A > Shelf 1")
        self.assertEqual(preview.rows[1].parent, "Freezer A > Shelf 1 > Rack Path")

        result = bulk_import_service.commit_box_import(
            self.session,
            BulkBoxImportCommitInput(raw_payload=raw_payload),
            self.user,
        )

        self.assertEqual(result.imported_rows, 2)
        nodes = storage_service.list_all_nodes(self.session)
        shelf_box = next(node for node in nodes if node.node_type.value == "box" and node.name == "Box Path Shelf")
        rack_box = next(node for node in nodes if node.node_type.value == "box" and node.name == "Box Path Rack")
        self.assertEqual(shelf_box.parent_id, self.shelf.id)
        self.assertEqual(rack_box.parent_id, existing_rack.id)

    def test_bulk_box_import_rejects_invalid_parent_and_duplicate_box(self):
        raw_payload = "\n".join(
            [
                "parent,box,rows,cols,box_nickname,notes",
                "Freezer A,Box 30,2,2,,Parent too shallow",
                "Freezer A > Shelf 1,Box 1,0,2,,Duplicate and invalid dimensions",
            ]
        )

        preview = bulk_import_service.preview_box_import(self.session, raw_payload)
        self.assertEqual(preview.valid_rows, 0)
        self.assertIn("Parent path was not found", preview.rows[0].errors)
        self.assertIn("Box name already exists", preview.rows[1].errors)
        self.assertIn("Rows must be a positive whole number", preview.rows[1].errors)

    def test_box_excel_template_contains_validations_and_roundtrips(self):
        workbook_bytes = bulk_import_service.box_template_xlsx(
            parent_paths=["Freezer A > Shelf 1", "Freezer A > Shelf 1 > Rack 1"],
        )
        workbook = load_workbook(filename=BytesIO(workbook_bytes))
        sheet = workbook["Box Import"]
        lookup_sheet = workbook["_lists"]

        self.assertEqual(workbook.sheetnames, ["Box Import", "_lists"])
        self.assertEqual(lookup_sheet.sheet_state, "hidden")
        self.assertEqual(sheet["A1"].value, "parent")
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertIsNotNone(sheet["A1"].comment)
        self.assertIn("existing shelf or rack path", sheet["A1"].comment.text)
        self.assertEqual(lookup_sheet["A2"].value, "Freezer A > Shelf 1")
        self.assertEqual(lookup_sheet["A3"].value, "Freezer A > Shelf 1 > Rack 1")

        validations = list(sheet.data_validations.dataValidation)
        validation_formulas = {validation.formula1 for validation in validations}
        self.assertIn("'_lists'!$A$2:$A$3", validation_formulas)
        self.assertIn("1", validation_formulas)

        sheet["A3"] = "Freezer A > Shelf 1"
        sheet["B3"] = "Box 99"
        sheet["C3"] = 2
        sheet["D3"] = 3
        buffer = BytesIO()
        workbook.save(buffer)

        raw_payload = bulk_import_service.box_workbook_to_csv(buffer.getvalue())
        self.assertIn("parent,box,rows,cols,box_nickname,notes", raw_payload)
        self.assertIn("Freezer A > Shelf 1,Box 99,2,3", raw_payload)


if __name__ == "__main__":
    unittest.main()
