from __future__ import annotations

import csv
import io
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path

from sqlalchemy.orm import Session

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.domain import models
from app.repositories import samples as sample_repository
from app.services import samples as sample_service
from app.services import storage as storage_service
from app.schemas import (
    BoxCreateInput,
    BulkBoxImportCommitInput,
    BulkBoxImportCommitResult,
    BulkBoxImportPreview,
    BulkBoxImportRow,
    BulkSampleImportCommitInput,
    BulkSampleImportCommitResult,
    BulkSampleImportPreview,
    BulkSampleImportRow,
    PlaceSampleInput,
    SampleCreateInput,
    StorageNodeCreate,
)

SAMPLE_HEADERS = [
    "sample_id",
    "sample_type",
    "study",
    "visit",
    "timepoint",
    "aliquot",
    "hemolysis",
    "study_role",
    "volume",
    "volume_units",
    "thaw_count",
    "notes",
    "collection_at",
    "placement_mode",
    "box",
    "position",
    "placement_group",
    "placement_offset",
]
BOX_HEADERS = [
    "parent",
    "box",
    "rows",
    "cols",
    "notes",
]
DATE_FORMAT = "%m/%d/%y %H:%M"
VALID_STUDY_ROLES = {"current", "retired"}
VALID_PLACEMENT_MODES = {"specific", "next_empty", "unplaced", "grouped", "offset"}
HEMOLYSIS_ALLOWED_VALUES = [f"{value:g}" for value in [1 + (index * 0.5) for index in range(13)]]
HEMOLYSIS_VALIDATION_FORMULA = '"' + ",".join(HEMOLYSIS_ALLOWED_VALUES) + '"'
SAMPLE_ENTRY_ROWS = 500
BOX_ENTRY_ROWS = 500
PATH_SEPARATOR = " > "
SAMPLE_TEMPLATE_ASSET_PATH = Path(__file__).resolve().parents[1] / "assets" / "sample-import-template.xlsx"
SAMPLE_TEMPLATE_WIDTHS = {
    "A": 11.140625,
    "B": 13.42578125,
    "C": 8.42578125,
    "D": 7.0,
    "E": 10.85546875,
    "F": 8.0,
    "G": 10.42578125,
    "H": 11.28515625,
    "I": 8.5703125,
    "J": 13.85546875,
    "K": 12.28515625,
    "L": 11.0,
    "M": 13.5703125,
    "N": 18.5703125,
    "O": 24.140625,
    "P": 9.0,
    "Q": 17.7109375,
    "R": 18.0,
}
SAMPLE_CENTERED_ENTRY_COLUMNS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "M", "N", "O", "P", "Q", "R"]
SAMPLE_HEADER_COMMENTS = {
    "sample_id": "Required.\nEnter the sample identifier for this row.\nThis is part of the sample's identity and is checked for duplicates together with sample_type, visit, timepoint, and aliquot.\nEx: IAS028, NRS28",
    "sample_type": "Required.\nEnter the configured sample type exactly as listed in the system.\nThis tells the importer what kind of sample is being created.\nEx: Plasma, Serum, PBMC",
    "study": "Optional.\nEnter the study name if this sample should be linked to a study.\nLeave blank if no study should be assigned.\nEx: IAS, NRS",
    "visit": "Optional.\nNumber input only.\nUse this when the sample belongs to a specific visit number.",
    "timepoint": "Optional.\nNumber input only.\nUse this to identify when within the visit or protocol the sample was collected.",
    "aliquot": "Optional.\nNatural number input only.\nUse this to distinguish multiple aliquots of the same sample set.",
    "hemolysis": "Optional.\nNumbers between 1-7 only (can use 0.5 increments).\nEnter the hemolysis classification if known using the CDC hemolysis palette; 1 is the lowest classification and 7 is the highest.\nReference: https://www.cdc.gov/vector-borne-diseases/php/laboratories/reference-tool-for-hemolysis-status.html",
    "study_role": "Optional.\nAllowed values: current or retired.\nUse this to mark whether the sample is the current preferred study sample or an older retained one.\nBlank defaults to current.",
    "volume": "Optional.\nNumber only.\nEnter the current sample volume if it is known.\nUse the next column for units.",
    "volume_units": "Optional.\nAllowed values: mL or uL.\nUse this with volume.\nBlank defaults to mL.",
    "thaw_count": "Optional.\nWhole numbers only.\nEnter how many times the sample has already been thawed.\nBlank defaults to 0.",
    "notes": "Optional.\nFree text.\nUse this for any useful sample-specific note you want saved with the record.",
    "collection_at": "Required.\nEnter the collection date and time using MM/DD/YY HH:MM.\nThis is when the sample was collected.",
    "placement_mode": "Required.\nChooses how each sample will be placed.\nUse specific for most rows; the other modes are for batch placement patterns (details below).\nSpecific: When the exact freezer location is known; requires box + position.\nnext_empty: When the sample should go into the next open slot in a box; requires box only and blank position.\nUnplaced: When creating a sample without a freezer location; leave box and position blank.\nGrouped: When related samples should use the same position across different boxes; requires the same placement_group on all related rows and one anchor row with box + position.\nOffset: When related samples should stay in the same box but shift positions from an anchor; requires the same placement_group on all related rows, one anchor row with box + position, and placement_offset on non-anchor rows.",
    "box": "Required only for placement modes that use a freezer box.\nEnter the existing box name exactly as it appears in the system.\nLeave blank for unplaced.",
    "position": "Required for specific, and used only on the anchor row for grouped or offset.\nLeave blank for next_empty, unplaced, and non-anchor grouped/offset rows.",
    "placement_group": "Required for grouped and offset.\nEnter a whole-number group ID so related rows are treated as one placement set.",
    "placement_offset": "Required only for non-anchor offset rows.\nEnter a whole number 0 or greater to shift placement forward from the anchor position in box scan order.",
}
BOX_TEMPLATE_WIDTHS = {
    "A": 32.0,
    "B": 18.0,
    "C": 8.0,
    "D": 8.0,
    "E": 28.0,
}
BOX_CENTERED_ENTRY_COLUMNS = ["A", "B", "C", "D"]
BOX_HEADER_COMMENTS = {
    "parent": "Required.\nChoose an existing shelf or rack path from the dropdown.\nExamples: Freezer A > Shelf 1 or Freezer A > Shelf 1 > Rack 2",
    "box": "Required.\nThe new unique box name to create.",
    "rows": "Required.\nNumber of box rows. Whole numbers only and must be greater than zero.",
    "cols": "Required.\nNumber of box columns. Whole numbers only and must be greater than zero.",
    "notes": "Optional.\nFree text notes to store on the box.",
}


class BulkImportError(Exception):
    pass


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[dict[str, str | None]]
    global_errors: list[str]


@dataclass
class PositionPlan:
    box_id: int
    box_name: str
    box_lookup_names: set[str]
    positions_by_label: dict[str, models.StoragePosition]
    occupied_labels: set[str]


def sample_template_xlsx(
    sample_types: list[str] | None = None,
    studies: list[str] | None = None,
    boxes: list[str] | None = None,
) -> bytes:
    workbook = load_workbook(filename=SAMPLE_TEMPLATE_ASSET_PATH)
    if "Placement Help" in workbook.sheetnames:
        del workbook["Placement Help"]
    sheet = workbook["Sample Import"]
    lookup_sheet = workbook["_lists"] if "_lists" in workbook.sheetnames else workbook.create_sheet("_lists")
    lookup_sheet.sheet_state = "hidden"
    _ensure_sample_template_header_comments(sheet)
    for column_letter, width in SAMPLE_TEMPLATE_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width

    sample_types = [value for value in (sample_types or []) if value]
    studies = [value for value in (studies or []) if value]
    boxes = [value for value in (boxes or []) if value]

    for row in lookup_sheet.iter_rows():
        for cell in row:
            cell.value = None

    lookup_sheet["A1"] = "sample_types"
    for index, value in enumerate(sample_types, start=2):
        lookup_sheet.cell(row=index, column=1, value=value)

    lookup_sheet["B1"] = "studies"
    for index, value in enumerate(studies, start=2):
        lookup_sheet.cell(row=index, column=2, value=value)

    lookup_sheet["C1"] = "boxes"
    for index, value in enumerate(boxes, start=2):
        lookup_sheet.cell(row=index, column=3, value=value)

    sheet.data_validations.dataValidation = []

    study_role_validation = DataValidation(
        type="list",
        formula1='"current,retired"',
        allow_blank=True,
    )
    study_role_validation.error = "Use current or retired."
    sheet.add_data_validation(study_role_validation)
    study_role_validation.add("H2:H500")

    hemolysis_validation = DataValidation(
        type="list",
        formula1=HEMOLYSIS_VALIDATION_FORMULA,
        allow_blank=True,
    )
    hemolysis_validation.error = "Use a hemolysis value from 1 to 7 in 0.5 increments."
    sheet.add_data_validation(hemolysis_validation)
    hemolysis_validation.add("G2:G500")

    volume_units_validation = DataValidation(
        type="list",
        formula1='"mL,uL"',
        allow_blank=True,
    )
    volume_units_validation.error = "Use mL or uL."
    sheet.add_data_validation(volume_units_validation)
    volume_units_validation.add("J2:J500")

    placement_mode_validation = DataValidation(
        type="list",
        formula1='"specific,next_empty,unplaced,grouped,offset"',
        allow_blank=False,
    )
    placement_mode_validation.error = "Use specific, next_empty, unplaced, grouped, or offset."
    sheet.add_data_validation(placement_mode_validation)
    placement_mode_validation.add("N2:N500")

    placement_group_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="1",
        allow_blank=True,
    )
    placement_group_validation.error = "Use a whole number greater than or equal to 1."
    sheet.add_data_validation(placement_group_validation)
    placement_group_validation.add("Q2:Q500")

    placement_offset_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    placement_offset_validation.error = "Use a whole number greater than or equal to 0."
    sheet.add_data_validation(placement_offset_validation)
    placement_offset_validation.add("R2:R500")

    if sample_types:
        sample_type_formula = f"'_lists'!$A$2:$A${len(sample_types) + 1}"
        sample_type_validation = DataValidation(type="list", formula1=sample_type_formula, allow_blank=False)
        sample_type_validation.error = "Choose a sample type from the configured list."
        sheet.add_data_validation(sample_type_validation)
        sample_type_validation.add(f"B2:B{SAMPLE_ENTRY_ROWS}")

    if studies:
        study_formula = f"'_lists'!$B$2:$B${len(studies) + 1}"
        study_validation = DataValidation(type="list", formula1=study_formula, allow_blank=True)
        study_validation.error = "Choose a study from the configured list."
        sheet.add_data_validation(study_validation)
        study_validation.add(f"C2:C{SAMPLE_ENTRY_ROWS}")

    if boxes:
        box_formula = f"'_lists'!$C$2:$C${len(boxes) + 1}"
        box_validation = DataValidation(type="list", formula1=box_formula, allow_blank=True)
        box_validation.error = "Choose a box from the configured list."
        sheet.add_data_validation(box_validation)
        box_validation.add(f"O2:O{SAMPLE_ENTRY_ROWS}")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ensure_sample_template_header_comments(sheet) -> None:
    for index, header in enumerate(SAMPLE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = header
        comment_text = SAMPLE_HEADER_COMMENTS.get(header)
        if comment_text:
            cell.comment = Comment(comment_text, "Sample Storage")
            _autosize_comment(cell.comment)
        else:
            cell.comment = None


def _autosize_comment(comment: Comment) -> None:
    lines = (comment.text or "").splitlines() or [""]
    longest_line = max(len(line) for line in lines)
    width = min(max(220, int(longest_line * 7.2) + 28), 560)
    chars_per_line = max(24, int((width - 28) / 7.2))
    visual_line_count = sum(max(1, math.ceil(len(line) / chars_per_line)) for line in lines)
    height = min(max(90, visual_line_count * 18 + 28), 420)
    comment.width = width
    comment.height = height
def workbook_to_csv(file_bytes: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    max_col = 0
    for row in rows:
        for index, value in enumerate(row, start=1):
            if value not in (None, ""):
                max_col = max(max_col, index)
    if max_col == 0:
        return ""

    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        cleaned_row = [_xlsx_value_to_text(value) for value in row[:max_col]]
        if not any(cell for cell in cleaned_row):
            continue
        writer.writerow(cleaned_row)
    return output.getvalue()


def sample_workbook_to_csv(file_bytes: bytes) -> str:
    return workbook_to_csv(file_bytes)


def box_template_csv() -> str:
    return ",".join(BOX_HEADERS) + "\n"


def box_template_xlsx(
    parent_paths: list[str] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Box Import"
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="1F5C4B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    centered_entry_alignment = Alignment(horizontal="center", vertical="center")
    notes_entry_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for index, header in enumerate(BOX_HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.comment = Comment(BOX_HEADER_COMMENTS.get(header, ""), "Sample Storage")
        _autosize_comment(cell.comment)
        sheet.column_dimensions[cell.column_letter].width = BOX_TEMPLATE_WIDTHS.get(cell.column_letter, 18)

    for row_index in range(2, BOX_ENTRY_ROWS + 1):
        for column_letter in BOX_CENTERED_ENTRY_COLUMNS:
            sheet[f"{column_letter}{row_index}"].alignment = centered_entry_alignment
        for column_letter in ("E",):
            sheet[f"{column_letter}{row_index}"].alignment = notes_entry_alignment

    lookup_sheet = workbook.create_sheet("_lists")
    lookup_sheet.sheet_state = "hidden"

    parent_paths = [value for value in (parent_paths or []) if value]

    lookup_sheet["A1"] = "parent_paths"
    for index, value in enumerate(parent_paths, start=2):
        lookup_sheet.cell(row=index, column=1, value=value)

    if parent_paths:
        parent_formula = f"'_lists'!$A$2:$A${len(parent_paths) + 1}"
        parent_validation = DataValidation(type="list", formula1=parent_formula, allow_blank=False)
        parent_validation.error = "Choose an existing shelf or rack path from the configured list."
        sheet.add_data_validation(parent_validation)
        parent_validation.add(f"A2:A{BOX_ENTRY_ROWS}")

    for column_letter in ("C", "D"):
        dimension_validation = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1="1",
            allow_blank=False,
        )
        dimension_validation.error = "Use a whole number greater than or equal to 1."
        sheet.add_data_validation(dimension_validation)
        dimension_validation.add(f"{column_letter}2:{column_letter}{BOX_ENTRY_ROWS}")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def box_workbook_to_csv(file_bytes: bytes) -> str:
    return workbook_to_csv(file_bytes)


def preview_sample_import(db: Session, raw_payload: str, target_box_id: int | None = None) -> BulkSampleImportPreview:
    table = _parse_table(raw_payload, SAMPLE_HEADERS, {"sample_id", "placement_mode"})
    preview = BulkSampleImportPreview(
        raw_payload=raw_payload,
        headers=table.headers,
        target_box_id=target_box_id,
        global_errors=list(table.global_errors),
    )

    sample_type_map = {sample_type.name.lower(): sample_type for sample_type in sample_repository.list_sample_types(db)}
    study_map = _study_lookup_map(sample_repository.list_studies(db))
    existing_sample_identity_keys = sample_repository.list_sample_identity_keys(db)
    box_plan_lookup = _build_box_plan_lookup(db, target_box_id, preview.global_errors)

    preview.rows = [
        BulkSampleImportRow(
            row_number=index + 2,
            sample_id=row.get("sample_id"),
            sample_type=row.get("sample_type"),
            study=row.get("study"),
            visit=row.get("visit"),
            timepoint=row.get("timepoint"),
            aliquot=row.get("aliquot"),
            hemolysis=row.get("hemolysis"),
            study_role=row.get("study_role"),
            volume=row.get("volume"),
            volume_units=row.get("volume_units"),
            thaw_count=row.get("thaw_count"),
            notes=row.get("notes"),
            collection_at=row.get("collection_at"),
            box=row.get("box"),
            position=row.get("position"),
            placement_mode=row.get("placement_mode"),
            placement_group=row.get("placement_group"),
            placement_offset=row.get("placement_offset"),
        )
        for index, row in enumerate(table.rows)
    ]
    duplicate_identities_in_file = _duplicate_sample_row_identities(preview.rows, sample_type_map)

    if preview.global_errors:
        for row in preview.rows:
            row.errors.extend(preview.global_errors)
            row.status = "invalid"
        _finalize_preview_counts(preview)
        return preview

    explicit_positions_by_box: dict[int, set[str]] = {}
    next_empty_rows_by_box: dict[int, list[BulkSampleImportRow]] = {}
    grouped_rows_by_group: dict[int, list[BulkSampleImportRow]] = {}
    offset_rows_by_group: dict[int, list[BulkSampleImportRow]] = {}
    for row in preview.rows:
        _validate_sample_row(
            row,
            sample_type_map=sample_type_map,
            study_map=study_map,
            existing_sample_identity_keys=existing_sample_identity_keys,
            duplicate_identities_in_file=duplicate_identities_in_file,
            box_plan_lookup=box_plan_lookup,
            explicit_positions_by_box=explicit_positions_by_box,
            default_target_box_id=target_box_id,
        )
        resolved_plan = _resolve_plan_for_row(box_plan_lookup, row.box, target_box_id)
        if not row.valid:
            continue
        if row.placement_mode == "next_empty" and resolved_plan is not None:
            next_empty_rows_by_box.setdefault(resolved_plan.box_id, []).append(row)
        elif row.placement_mode == "grouped":
            group_id = _parse_int(row.placement_group)
            if group_id is not None:
                grouped_rows_by_group.setdefault(group_id, []).append(row)
        elif row.placement_mode == "offset":
            group_id = _parse_int(row.placement_group)
            if group_id is not None:
                offset_rows_by_group.setdefault(group_id, []).append(row)

    explicit_positions_by_box = _reserved_positions(preview.rows, box_plan_lookup)
    _resolve_grouped_rows(grouped_rows_by_group, box_plan_lookup, explicit_positions_by_box, target_box_id)
    explicit_positions_by_box = _reserved_positions(preview.rows, box_plan_lookup)
    _resolve_offset_rows(offset_rows_by_group, box_plan_lookup, explicit_positions_by_box, target_box_id)
    explicit_positions_by_box = _reserved_positions(preview.rows, box_plan_lookup)
    _resolve_next_empty_rows(next_empty_rows_by_box, box_plan_lookup, explicit_positions_by_box)
    _finalize_preview_counts(preview)
    return preview


def commit_sample_import(
    db: Session,
    data: BulkSampleImportCommitInput,
    user: models.User | None,
) -> BulkSampleImportCommitResult:
    preview = preview_sample_import(db, data.raw_payload, data.target_box_id)
    result = BulkSampleImportCommitResult(target_box_id=data.target_box_id, global_errors=list(preview.global_errors))
    if preview.global_errors:
        result.rows = [row.model_copy(update={"status": "skipped"}) for row in preview.rows]
        result.skipped_rows = len(result.rows)
        return result

    box_plan_lookup = _build_box_plan_lookup(db, data.target_box_id, [])
    box_name_lookup = {plan.box_name.lower(): plan for plan in box_plan_lookup.values()}
    sample_type_map = {sample_type.name.lower(): sample_type for sample_type in sample_repository.list_sample_types(db)}
    study_map = _study_lookup_map(sample_repository.list_studies(db))

    for row in preview.rows:
        row_copy = row.model_copy(deep=True)
        if not row.valid:
            row_copy.status = "skipped"
            result.rows.append(row_copy)
            result.skipped_rows += 1
            continue
        try:
            create_input = SampleCreateInput(
                sample_id=row.sample_id or "",
                sample_type_id=sample_type_map[row.sample_type.lower()].id if row.sample_type else None,
                study_id=study_map[row.study.lower()].id if row.study else None,
                visit_label=row.visit,
                timepoint_label=row.timepoint,
                aliquot_number=_parse_int(row.aliquot),
                hemolysis_classification=_parse_float(row.hemolysis),
                study_role=(row.study_role or "current"),
                volume=_parse_float(row.volume),
                volume_units=row.volume_units or "mL",
                thaw_count=_parse_int(row.thaw_count) or 0,
                notes=row.notes,
                collection_at=_parse_datetime(row.collection_at),
            )
            sample = sample_service.create_sample(db, create_input, user, commit=False)
            if row.assigned_box_name and row.assigned_position:
                position_plan = box_name_lookup.get(row.assigned_box_name.lower())
                if position_plan is None:
                    raise BulkImportError("Assigned box was not found during import")
                position_id = position_plan.positions_by_label[row.assigned_position].id
                sample_service.place_sample(db, sample.id, PlaceSampleInput(position_id=position_id), user, commit=False)
            db.commit()
            row_copy.status = "imported"
            result.rows.append(row_copy)
            result.imported_rows += 1
        except Exception as exc:
            db.rollback()
            row_copy.status = "failed"
            row_copy.errors.append(str(exc))
            result.rows.append(row_copy)
            result.failed_rows += 1
    return result


def preview_box_import(db: Session, raw_payload: str) -> BulkBoxImportPreview:
    table = _parse_table(raw_payload, BOX_HEADERS, {"parent", "box", "rows", "cols"})
    preview = BulkBoxImportPreview(raw_payload=raw_payload, headers=table.headers, global_errors=list(table.global_errors))
    all_nodes = storage_service.list_all_nodes(db)
    parent_lookup = _box_parent_lookup(all_nodes)
    preview.rows = [
        BulkBoxImportRow(
            row_number=index + 2,
            parent=row.get("parent"),
            box=row.get("box"),
            rows=row.get("rows"),
            cols=row.get("cols"),
            notes=row.get("notes"),
        )
        for index, row in enumerate(table.rows)
    ]
    if preview.global_errors:
        for row in preview.rows:
            row.errors.extend(preview.global_errors)
        _finalize_box_counts(preview)
        return preview

    box_duplicates = _duplicate_keys(table.rows, "box")
    box_name_conflicts = _existing_box_names(all_nodes)

    for row in preview.rows:
        _validate_box_row(
            row,
            parent_lookup=parent_lookup,
            box_duplicates=box_duplicates,
            existing_box_names=box_name_conflicts,
        )
    _finalize_box_counts(preview)
    return preview


def commit_box_import(
    db: Session,
    data: BulkBoxImportCommitInput,
    user: models.User | None,
) -> BulkBoxImportCommitResult:
    preview = preview_box_import(db, data.raw_payload)
    result = BulkBoxImportCommitResult(global_errors=list(preview.global_errors))
    if preview.global_errors:
        result.rows = [row.model_copy(update={"status": "skipped"}) for row in preview.rows]
        result.skipped_rows = len(result.rows)
        return result

    for row in preview.rows:
        row_copy = row.model_copy(deep=True)
        if not row.valid:
            row_copy.status = "skipped"
            result.rows.append(row_copy)
            result.skipped_rows += 1
            continue
        try:
            all_nodes = storage_service.list_all_nodes(db)
            parent = _resolve_box_parent(all_nodes, row.parent)
            if parent is None:
                raise BulkImportError("Storage path no longer exists")
            box = storage_service.create_storage_node(
                db,
                StorageNodeCreate(
                    name=row.box or "",
                    notes=row.notes,
                    node_type="box",
                    parent_id=parent.id,
                ),
                user,
                commit=False,
            )
            storage_service.create_box_positions(
                db,
                BoxCreateInput(box_id=box.id, rows=_parse_int(row.rows) or 0, cols=_parse_int(row.cols) or 0),
                user,
                commit=False,
            )
            db.commit()
            row_copy.status = "imported"
            result.rows.append(row_copy)
            result.imported_rows += 1
        except Exception as exc:
            db.rollback()
            row_copy.status = "failed"
            row_copy.errors.append(str(exc))
            result.rows.append(row_copy)
            result.failed_rows += 1
    return result


def _parse_table(raw_payload: str, allowed_headers: list[str], required_headers: set[str]) -> ParsedTable:
    payload = (raw_payload or "").replace("\ufeff", "").strip()
    if not payload:
        return ParsedTable(headers=[], rows=[], global_errors=["No import data provided"])
    first_line = payload.splitlines()[0] if payload.splitlines() else ""
    delimiter = "\t" if "\t" in first_line else ","
    reader = csv.DictReader(io.StringIO(payload), delimiter=delimiter)
    headers = [header.strip() for header in (reader.fieldnames or []) if header and header.strip()]
    global_errors: list[str] = []
    missing_headers = sorted(required_headers.difference(headers))
    unknown_headers = sorted(set(headers).difference(allowed_headers))
    if missing_headers:
        global_errors.append(f"Missing required columns: {', '.join(missing_headers)}")
    if unknown_headers:
        global_errors.append(f"Unknown columns: {', '.join(unknown_headers)}")
    rows: list[dict[str, str | None]] = []
    for row in reader:
        cleaned = {key.strip(): _clean_value(value) for key, value in row.items() if key is not None}
        if not any(value for value in cleaned.values()):
            continue
        rows.append(cleaned)
    return ParsedTable(headers=headers, rows=rows, global_errors=global_errors)


def _validate_sample_row(
    row: BulkSampleImportRow,
    *,
    sample_type_map: dict[str, models.SampleType],
    study_map: dict[str, models.Study],
    existing_sample_identity_keys: set[tuple[str, int | None, str | None, str | None, int | None]],
    duplicate_identities_in_file: set[tuple[str, int | None, str | None, str | None, int | None]],
    box_plan_lookup: dict[int, PositionPlan],
    explicit_positions_by_box: dict[int, set[str]],
    default_target_box_id: int | None,
) -> None:
    row.placement_mode = _clean_value(row.placement_mode)
    if row.placement_mode:
        row.placement_mode = row.placement_mode.lower()
    row.placement_group = _clean_value(row.placement_group)
    row.placement_offset = _clean_value(row.placement_offset)
    sample_id = (row.sample_id or "").strip()
    if not sample_id:
        row.errors.append("Sample ID is required")

    if not row.sample_type:
        row.errors.append("Sample type is required")
    elif row.sample_type.lower() not in sample_type_map:
        row.errors.append("Sample type was not found")
    if row.study and row.study.lower() not in study_map:
        row.errors.append("Study was not found")
    if row.visit and not str(row.visit).strip().isdigit():
        row.errors.append("Visit must contain numbers only")
    if row.timepoint and not str(row.timepoint).strip().isdigit():
        row.errors.append("Timepoint must contain numbers only")
    if row.study_role and row.study_role not in VALID_STUDY_ROLES:
        row.errors.append("Study role must be current or retired")
    if row.aliquot and _parse_int(row.aliquot) is None:
        row.errors.append("Aliquot must be a whole number")
    if row.hemolysis:
        hemolysis_value = _parse_float(row.hemolysis)
        if hemolysis_value is None:
            row.errors.append("Hemolysis must be between 1 and 7 in 0.5 increments")
        else:
            try:
                sample_service._normalize_hemolysis(hemolysis_value)
            except sample_service.SampleError as exc:
                row.errors.append(str(exc))
    if row.thaw_count and (_parse_int(row.thaw_count) is None or (_parse_int(row.thaw_count) or 0) < 0):
        row.errors.append("Thaw count must be zero or greater")
    if row.volume and (_parse_float(row.volume) is None or (_parse_float(row.volume) or 0) < 0):
        row.errors.append("Volume must be zero or greater")
    if row.volume_units and row.volume_units not in {"mL", "uL"}:
        row.errors.append("Volume units must be mL or uL")
    if not row.collection_at:
        row.errors.append("Collection date is required")
    elif _parse_datetime(row.collection_at) is None:
        row.errors.append(f"Collection date must use {DATE_FORMAT}")
    if not row.placement_mode:
        row.errors.append("Placement mode is required")
    elif row.placement_mode not in VALID_PLACEMENT_MODES:
        row.errors.append("Placement mode must be specific, next_empty, unplaced, grouped, or offset")
    placement_group_value = _parse_int(row.placement_group)
    if row.placement_group and placement_group_value is None:
        row.errors.append("Placement group must be a whole number")
    elif placement_group_value is not None and placement_group_value <= 0:
        row.errors.append("Placement group must be greater than zero")
    if row.placement_group and row.placement_mode not in {"grouped", "offset"}:
        row.errors.append("Placement group can only be used when placement mode is grouped or offset")
    placement_offset_value = _parse_int(row.placement_offset)
    if row.placement_offset and placement_offset_value is None:
        row.errors.append("Placement offset must be a whole number")
    elif placement_offset_value is not None and placement_offset_value < 0:
        row.errors.append("Placement offset must be zero or greater")
    if row.placement_offset and row.placement_mode != "offset":
        row.errors.append("Placement offset can only be used when placement mode is offset")

    sample_type = sample_type_map.get((row.sample_type or "").lower()) if row.sample_type else None
    identity_key = None
    if sample_id and sample_type is not None:
        identity_key = sample_repository.build_identity_key(
            sample_id,
            sample_type.id,
            row.visit,
            row.timepoint,
            _parse_int(row.aliquot),
        )
        if identity_key in existing_sample_identity_keys:
            row.errors.append("This ID, type, visit, timepoint, and aliquot combination already exists")
        elif identity_key in duplicate_identities_in_file:
            row.errors.append("This ID, type, visit, timepoint, and aliquot combination is duplicated in this import")

    position_plan = _resolve_plan_for_row(box_plan_lookup, row.box, default_target_box_id)
    if row.box and position_plan is None:
        row.errors.append("Box was not found")
    elif position_plan is not None and row.placement_mode != "unplaced":
        row.assigned_box_name = position_plan.box_name

    if row.placement_mode == "specific":
        if position_plan is None:
            row.errors.append("A box is required when placement mode is specific")
        if not row.position:
            row.errors.append("Position is required when placement mode is specific")
        elif position_plan is not None:
            label = row.position.upper()
            explicit_positions = explicit_positions_by_box.setdefault(position_plan.box_id, set())
            if label not in position_plan.positions_by_label:
                row.errors.append("Position does not exist in the resolved box")
            elif label in position_plan.occupied_labels:
                row.errors.append("Position is already occupied")
            elif label in explicit_positions:
                row.errors.append("Position is duplicated in this import for the same box")
            else:
                row.assigned_box_name = position_plan.box_name
                row.assigned_position = label
                explicit_positions.add(label)
    elif row.placement_mode == "unplaced":
        if row.box:
            row.errors.append("Box must be blank when placement mode is unplaced")
        if row.position:
            row.errors.append("Position must be blank when placement mode is unplaced")
    elif row.placement_mode == "next_empty":
        if position_plan is None:
            row.errors.append("A box is required when placement mode is next_empty")
        if row.position:
            row.errors.append("Position must be blank when placement mode is next_empty")
    elif row.placement_mode == "grouped":
        if position_plan is None:
            row.errors.append("A box is required when placement mode is grouped")
        if placement_group_value is None:
            row.errors.append("Placement group is required when placement mode is grouped")
        if row.position:
            if position_plan is None:
                row.errors.append("Position can only be used when a box is resolved")
            else:
                label = row.position.upper()
                explicit_positions = explicit_positions_by_box.setdefault(position_plan.box_id, set())
                if label not in position_plan.positions_by_label:
                    row.errors.append("Position does not exist in the resolved box")
                elif label in position_plan.occupied_labels:
                    row.errors.append("Position is already occupied")
                elif label in explicit_positions:
                    row.errors.append("Position is duplicated in this import for the same box")
                else:
                    row.assigned_box_name = position_plan.box_name
                    row.assigned_position = label
                    explicit_positions.add(label)
    elif row.placement_mode == "offset":
        if placement_group_value is None:
            row.errors.append("Placement group is required when placement mode is offset")
        if row.position:
            if position_plan is None:
                row.errors.append("Position can only be used when a box is resolved")
            else:
                label = row.position.upper()
                explicit_positions = explicit_positions_by_box.setdefault(position_plan.box_id, set())
                if label not in position_plan.positions_by_label:
                    row.errors.append("Position does not exist in the resolved box")
                elif label in position_plan.occupied_labels:
                    row.errors.append("Position is already occupied")
                elif label in explicit_positions:
                    row.errors.append("Position is duplicated in this import for the same box")
                else:
                    row.assigned_box_name = position_plan.box_name
                    row.assigned_position = label
                    explicit_positions.add(label)
        elif placement_offset_value is None:
            row.errors.append("Placement offset is required for non-anchor offset rows")

    row.valid = not row.errors
    row.status = "valid" if row.valid else "invalid"


def _resolve_grouped_rows(
    grouped_rows_by_group: dict[int, list[BulkSampleImportRow]],
    box_plan_lookup: dict[int, PositionPlan],
    explicit_positions_by_box: dict[int, set[str]],
    default_target_box_id: int | None,
) -> None:
    for group_id, rows in grouped_rows_by_group.items():
        anchor_rows = [row for row in rows if row.position]
        if len(anchor_rows) != 1:
            for row in rows:
                row.errors.append("Grouped placement requires exactly one anchor row with a position")
                row.valid = False
                row.status = "invalid"
            continue
        anchor = anchor_rows[0]
        anchor_plan = _resolve_plan_for_row(box_plan_lookup, anchor.box, default_target_box_id)
        if anchor_plan is None:
            for row in rows:
                row.errors.append("Grouped placement anchor box could not be resolved")
                row.valid = False
                row.status = "invalid"
            continue
        anchor_label = anchor.position.upper()
        seen_box_ids = {anchor_plan.box_id}
        for row in rows:
            if row is anchor or not row.valid:
                continue
            if row.position:
                row.errors.append("Only the anchor row may include a position for grouped placement")
                row.valid = False
                row.status = "invalid"
                continue
            position_plan = _resolve_plan_for_row(box_plan_lookup, row.box, default_target_box_id)
            if position_plan is None:
                row.errors.append("Grouped placement box could not be resolved")
                row.valid = False
                row.status = "invalid"
                continue
            if position_plan.box_id in seen_box_ids:
                row.errors.append(f"Placement group {group_id} cannot target the same box more than once")
                row.valid = False
                row.status = "invalid"
                continue
            if anchor_label not in position_plan.positions_by_label:
                row.errors.append(f"Position {anchor_label} does not exist in the grouped box")
                row.valid = False
                row.status = "invalid"
                continue
            explicit_positions = explicit_positions_by_box.setdefault(position_plan.box_id, set())
            if anchor_label in position_plan.occupied_labels:
                row.errors.append(f"Position {anchor_label} is already occupied in the grouped box")
                row.valid = False
                row.status = "invalid"
                continue
            if anchor_label in explicit_positions:
                row.errors.append(f"Position {anchor_label} is duplicated in this import for the grouped box")
                row.valid = False
                row.status = "invalid"
                continue
            row.assigned_box_name = position_plan.box_name
            row.assigned_position = anchor_label
            explicit_positions.add(anchor_label)
            seen_box_ids.add(position_plan.box_id)


def _resolve_offset_rows(
    offset_rows_by_group: dict[int, list[BulkSampleImportRow]],
    box_plan_lookup: dict[int, PositionPlan],
    explicit_positions_by_box: dict[int, set[str]],
    default_target_box_id: int | None,
) -> None:
    for group_id, rows in offset_rows_by_group.items():
        anchor_rows = [row for row in rows if row.position]
        if len(anchor_rows) != 1:
            for row in rows:
                row.errors.append("Offset placement requires exactly one anchor row with a position")
                row.valid = False
                row.status = "invalid"
            continue
        anchor = anchor_rows[0]
        anchor_plan = _resolve_plan_for_row(box_plan_lookup, anchor.box, default_target_box_id)
        if anchor_plan is None:
            for row in rows:
                row.errors.append("Offset placement anchor box could not be resolved")
                row.valid = False
                row.status = "invalid"
            continue
        anchor_offset = _parse_int(anchor.placement_offset)
        if anchor_offset not in (None, 0):
            for row in rows:
                row.errors.append("Offset placement anchor row must leave placement_offset blank or set it to 0")
                row.valid = False
                row.status = "invalid"
            continue
        anchor_label = anchor.position.upper()
        ordered_labels = _ordered_position_labels(anchor_plan)
        if anchor_label not in ordered_labels:
            for row in rows:
                row.errors.append("Offset placement anchor position was not found")
                row.valid = False
                row.status = "invalid"
            continue
        anchor_index = ordered_labels.index(anchor_label)
        for row in rows:
            if row is anchor or not row.valid:
                continue
            if row.position:
                row.errors.append("Only the anchor row may include a position for offset placement")
                row.valid = False
                row.status = "invalid"
                continue
            position_plan = _resolve_plan_for_row(box_plan_lookup, row.box, default_target_box_id)
            if row.box and position_plan is not None and position_plan.box_id != anchor_plan.box_id:
                row.errors.append("Offset placement rows must use the same box as the anchor")
                row.valid = False
                row.status = "invalid"
                continue
            offset_value = _parse_int(row.placement_offset)
            if offset_value is None:
                row.errors.append("Placement offset is required for non-anchor offset rows")
                row.valid = False
                row.status = "invalid"
                continue
            target_index = anchor_index + offset_value
            if target_index >= len(ordered_labels):
                row.errors.append("Placement offset extends past the end of the anchor box")
                row.valid = False
                row.status = "invalid"
                continue
            target_label = ordered_labels[target_index]
            explicit_positions = explicit_positions_by_box.setdefault(anchor_plan.box_id, set())
            if target_label in anchor_plan.occupied_labels:
                row.errors.append(f"Position {target_label} is already occupied in the offset box")
                row.valid = False
                row.status = "invalid"
                continue
            if target_label in explicit_positions:
                row.errors.append(f"Position {target_label} is duplicated in this import for the offset box")
                row.valid = False
                row.status = "invalid"
                continue
            row.assigned_box_name = anchor_plan.box_name
            row.assigned_position = target_label
            explicit_positions.add(target_label)


def _resolve_next_empty_rows(
    next_empty_rows_by_box: dict[int, list[BulkSampleImportRow]],
    box_plan_lookup: dict[int, PositionPlan],
    explicit_positions_by_box: dict[int, set[str]],
) -> None:
    for box_id, rows in next_empty_rows_by_box.items():
        position_plan = box_plan_lookup.get(box_id)
        if position_plan is None:
            continue
        available_labels = _available_labels(position_plan, explicit_positions_by_box.get(box_id, set()))
        for row in rows:
            if not available_labels:
                row.errors.append("No open positions remain for next_empty placement")
                row.valid = False
                row.status = "invalid"
                continue
            label = available_labels.pop(0)
            explicit_positions_by_box.setdefault(box_id, set()).add(label)
            row.assigned_box_name = position_plan.box_name
            row.assigned_position = label
            row.status = "valid"


def _build_box_plan_lookup(db: Session, target_box_id: int | None, global_errors: list[str]) -> dict[int, PositionPlan]:
    plans: dict[int, PositionPlan] = {}
    for box in storage_service.list_boxes(db):
        plan = _build_position_plan_for_box(db, box)
        if plan is not None:
            plans[box.id] = plan
    if target_box_id is not None and target_box_id not in plans:
        global_errors.append("Selected target box was not found")
    return plans


def _build_position_plan_for_box(db: Session, box_node: models.StorageNode) -> PositionPlan | None:
    box = storage_service.get_box_view(db, box_node.id)
    if box is None:
        return None
    position_map = {position.label.upper(): storage_service.get_position(db, position.id) for position in box.positions}
    if any(position is None for position in position_map.values()):
        return None
    positions = {label: position for label, position in position_map.items() if position is not None}
    occupied = {label for label, position in positions.items() if position.location is not None}
    lookup_names = {box.box_name.lower(), box_node.name.lower()}
    return PositionPlan(
        box_id=box.box_id,
        box_name=box.box_name,
        box_lookup_names=lookup_names,
        positions_by_label=positions,
        occupied_labels=occupied,
    )


def _validate_box_row(
    row: BulkBoxImportRow,
    *,
    parent_lookup: dict[str, models.StorageNode],
    box_duplicates: set[str],
    existing_box_names: set[str],
) -> None:
    row.parent = _clean_value(row.parent)
    parent = row.parent or ""
    box = (row.box or "").strip()
    if not parent:
        row.errors.append("Parent path is required")
    elif parent not in parent_lookup:
        row.errors.append("Parent path was not found")
    if not box:
        row.errors.append("Box is required")
    parsed_rows = _parse_int(row.rows)
    parsed_cols = _parse_int(row.cols)
    if row.rows is None or parsed_rows is None or parsed_rows <= 0:
        row.errors.append("Rows must be a positive whole number")
    if row.cols is None or parsed_cols is None or parsed_cols <= 0:
        row.errors.append("Cols must be a positive whole number")
    if box and box.lower() in box_duplicates:
        row.errors.append("Box name is duplicated in this import")
    if box and box in existing_box_names:
        row.errors.append("Box name already exists")

    row.valid = not row.errors
    row.status = "valid" if row.valid else "invalid"


def _box_parent_lookup(all_nodes: list[models.StorageNode]) -> dict[str, models.StorageNode]:
    return {
        PATH_SEPARATOR.join(node.path_names()): node
        for node in all_nodes
        if node.node_type in {models.StorageNodeType.shelf, models.StorageNodeType.rack}
    }


def _resolve_box_parent(all_nodes: list[models.StorageNode], parent_path: str | None) -> models.StorageNode | None:
    if not parent_path:
        return None
    return _box_parent_lookup(all_nodes).get(parent_path)


def _duplicate_keys(rows: list[dict[str, str | None]], key: str) -> set[str]:
    counts: dict[str, int] = {}
    for row in rows:
        value = (row.get(key) or "").strip().lower()
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return {value for value, count in counts.items() if count > 1}


def _duplicate_sample_row_identities(
    rows: list[BulkSampleImportRow],
    sample_type_map: dict[str, models.SampleType],
) -> set[tuple[str, int | None, str | None, str | None, int | None]]:
    counts: dict[tuple[str, int | None, str | None, str | None, int | None], int] = {}
    for row in rows:
        sample_id = (row.sample_id or "").strip()
        sample_type = sample_type_map.get((row.sample_type or "").lower()) if row.sample_type else None
        if not sample_id or sample_type is None:
            continue
        identity_key = sample_repository.build_identity_key(
            sample_id,
            sample_type.id,
            row.visit,
            row.timepoint,
            _parse_int(row.aliquot),
        )
        counts[identity_key] = counts.get(identity_key, 0) + 1
    return {identity_key for identity_key, count in counts.items() if count > 1}


def _existing_box_names(all_nodes: list[models.StorageNode]) -> set[str]:
    return {node.name for node in all_nodes if node.node_type == models.StorageNodeType.box}


def _study_lookup_map(studies: list[models.Study]) -> dict[str, models.Study]:
    lookup: dict[str, models.Study] = {}
    for study in studies:
        lookup[study.name.lower()] = study
    return lookup


def _resolve_plan_for_row(
    box_plan_lookup: dict[int, PositionPlan],
    row_box: str | None,
    default_target_box_id: int | None,
) -> PositionPlan | None:
    if row_box:
        normalized = row_box.strip().lower()
        for plan in box_plan_lookup.values():
            if normalized in plan.box_lookup_names:
                return plan
        return None
    if default_target_box_id is not None:
        return box_plan_lookup.get(default_target_box_id)
    return None


def _available_labels(position_plan: PositionPlan, reserved_labels: set[str]) -> list[str]:
    return [
        label
        for label in _ordered_position_labels(position_plan)
        if label not in position_plan.occupied_labels and label not in reserved_labels
    ]


def _ordered_position_labels(position_plan: PositionPlan) -> list[str]:
    return [
        label
        for label, _position in sorted(position_plan.positions_by_label.items(), key=lambda item: (item[1].row, item[1].col))
    ]


def _reserved_positions(
    rows: list[BulkSampleImportRow],
    box_plan_lookup: dict[int, PositionPlan],
) -> dict[int, set[str]]:
    reserved: dict[int, set[str]] = {}
    box_lookup = {plan.box_name.lower(): plan for plan in box_plan_lookup.values()}
    for row in rows:
        if not row.valid or not row.assigned_box_name or not row.assigned_position:
            continue
        plan = box_lookup.get(row.assigned_box_name.lower())
        if plan is None:
            continue
        reserved.setdefault(plan.box_id, set()).add(row.assigned_position)
    return reserved


def _finalize_preview_counts(preview: BulkSampleImportPreview) -> None:
    preview.total_rows = len(preview.rows)
    preview.valid_rows = sum(1 for row in preview.rows if row.valid)
    preview.invalid_rows = preview.total_rows - preview.valid_rows


def _finalize_box_counts(preview: BulkBoxImportPreview) -> None:
    preview.total_rows = len(preview.rows)
    preview.valid_rows = sum(1 for row in preview.rows if row.valid)
    preview.invalid_rows = preview.total_rows - preview.valid_rows


def _parse_int(value: str | None) -> int | None:
    if value is None or not str(value).strip():
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def _parse_float(value: str | None) -> float | None:
    if value is None or not str(value).strip():
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _parse_datetime(value: str | None) -> datetime | None:
    if value is None or not str(value).strip():
        return None
    try:
        return datetime.strptime(str(value).strip(), DATE_FORMAT)
    except ValueError:
        return None


def _clean_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _xlsx_value_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
