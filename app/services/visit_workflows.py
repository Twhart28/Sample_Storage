from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload, selectinload

from app.domain import models
from app.schemas import (
    BulkSampleImportCommitResult,
    BulkSampleImportPreview,
    PlaceSampleInput,
    SampleCreateInput,
    StudyWorkflowConfigInput,
    StudyWorkflowQuickLink,
    StudyWorkflowView,
    VisitSessionCompleteInput,
    VisitSessionCreateInput,
    VisitSessionNotesInput,
    VisitSessionSampleLink,
    VisitSessionView,
    VisitWorkbookCommitResult,
    VisitWorkbookPreview,
    VisitWorkflowListItem,
)
from app.services import bulk_imports as bulk_import_service
from app.services import samples as sample_service
from app.services import storage as storage_service

VISIT_SAMPLE_SHEET = "sample_Import"
STEP_TEMPLATE_DOWNLOADED = "template_downloaded_at"
STEP_BIOCHEM_OPENED = "biochem_lipid_opened_at"
STEP_DILUTION_OPENED = "dilution_calculator_opened_at"
STEP_WORKBOOK_UPLOADED = "workbook_uploaded_at"
STEP_SAMPLE_IMPORT_COMMITTED = "sample_import_committed_at"
STEP_SUMMARY_REVIEWED = "summary_reviewed_at"


class VisitWorkflowError(Exception):
    pass


def list_active_workflows(db: Session) -> list[VisitWorkflowListItem]:
    workflows = _list_workflows(db, active_only=True)
    return [
        VisitWorkflowListItem(
            study_id=workflow.study_id,
            study_name=workflow.study.name,
            workflow_id=workflow.id,
            label=workflow.label,
            description=workflow.description,
            is_active=workflow.is_active,
        )
        for workflow in workflows
    ]


def count_active_workflows(db: Session) -> int:
    return len(list_active_workflows(db))


def list_recent_sessions(
    db: Session,
    limit: int = 8,
    *,
    status: models.VisitSessionStatus | str | None = None,
    submitted_only: bool = False,
) -> list[VisitSessionView]:
    stmt = (
        select(models.VisitSession)
        .options(
            joinedload(models.VisitSession.study),
            joinedload(models.VisitSession.workflow),
            joinedload(models.VisitSession.operator),
            selectinload(models.VisitSession.samples)
            .joinedload(models.VisitSessionSample.sample)
            .joinedload(models.Sample.sample_type),
            selectinload(models.VisitSession.samples)
            .joinedload(models.VisitSessionSample.sample)
            .selectinload(models.Sample.location)
            .joinedload(models.SampleLocation.position),
        )
    )
    if status is not None:
        status_value = status if isinstance(status, models.VisitSessionStatus) else models.VisitSessionStatus(str(status))
        stmt = stmt.where(models.VisitSession.status == status_value)
    if submitted_only:
        stmt = stmt.where(models.VisitSession.status == models.VisitSessionStatus.completed)
        stmt = stmt.order_by(models.VisitSession.completed_at.desc().nullslast(), models.VisitSession.updated_at.desc())
    else:
        stmt = stmt.order_by(models.VisitSession.updated_at.desc())
    stmt = stmt.limit(limit)
    return [_build_session_view(session) for session in db.execute(stmt).unique().scalars().all()]


def list_recent_draft_sessions(db: Session, limit: int = 5) -> list[VisitSessionView]:
    return list_recent_sessions(db, limit=limit, status=models.VisitSessionStatus.draft)


def list_recent_submitted_sessions(db: Session, limit: int = 5) -> list[VisitSessionView]:
    return list_recent_sessions(db, limit=limit, submitted_only=True)


def get_workflow_for_study(db: Session, study_id: int) -> StudyWorkflowView:
    workflow = _get_or_create_workflow(db, study_id)
    db.refresh(workflow)
    return _build_workflow_view(workflow)


def save_workflow_config(
    db: Session,
    study_id: int,
    data: StudyWorkflowConfigInput,
    *,
    template_filename: str | None = None,
    template_bytes: bytes | None = None,
) -> StudyWorkflowView:
    workflow = _get_or_create_workflow(db, study_id)
    workflow.label = data.label.strip()
    workflow.description = _clean_text(data.description)
    workflow.is_active = bool(data.is_active)
    workflow.set_quick_links(_build_quick_links(data.quick_links))
    if template_bytes is not None:
        validated_bytes = _validate_visit_template_workbook(template_bytes)
        workflow.template_workbook_filename = _clean_text(template_filename) or "visit-template.xlsx"
        workflow.template_workbook_blob = validated_bytes
    workflow.updated_at = datetime.utcnow()
    db.add(workflow)
    db.commit()
    db.refresh(workflow)
    return _build_workflow_view(workflow)


def create_visit_session(db: Session, data: VisitSessionCreateInput, user: models.User | None) -> VisitSessionView:
    workflow = _get_active_workflow_for_study(db, data.study_id)
    participant_id = (data.participant_id or "").strip()
    if not participant_id:
        raise VisitWorkflowError("Participant ID is required")
    session = models.VisitSession(
        study_id=workflow.study_id,
        workflow_id=workflow.id,
        participant_id=participant_id,
        visit_date=data.visit_date,
        operator_user_id=user.id if user else None,
        status=models.VisitSessionStatus.draft,
    )
    session.set_step_status({})
    db.add(session)
    db.commit()
    db.refresh(session)
    return get_visit_session(db, session.id)


def get_visit_session(db: Session, session_id: int) -> VisitSessionView:
    session = _get_session(db, session_id)
    return _build_session_view(session)


def update_visit_notes(
    db: Session,
    session_id: int,
    data: VisitSessionNotesInput,
) -> VisitSessionView:
    session = _get_session(db, session_id)
    session.session_notes = _clean_text(data.notes)
    session.deviation_notes = None
    session.updated_at = datetime.utcnow()
    db.add(session)
    db.commit()
    db.refresh(session)
    return _build_session_view(session)


def mark_link_opened(db: Session, session_id: int, link_key: str) -> tuple[VisitSessionView, str]:
    session = _get_session(db, session_id)
    link = next((item for item in session.workflow.quick_links if item.get("key") == link_key), None)
    if link is None or not link.get("url"):
        raise VisitWorkflowError("Workflow link was not found")
    step_map = {
        "biochem_lipid": STEP_BIOCHEM_OPENED,
        "dilution_calculator": STEP_DILUTION_OPENED,
    }
    _mark_step(session, step_map.get(link_key, f"{link_key}_opened_at"))
    db.add(session)
    db.commit()
    db.refresh(session)
    return _build_session_view(session), str(link["url"])


def mark_summary_reviewed(db: Session, session_id: int) -> VisitSessionView:
    session = _get_session(db, session_id)
    _mark_step(session, STEP_SUMMARY_REVIEWED)
    db.add(session)
    db.commit()
    db.refresh(session)
    return _build_session_view(session)


def generate_visit_template_xlsx(db: Session, session_id: int) -> tuple[str, bytes]:
    session = _get_session(db, session_id)
    workflow = session.workflow
    if workflow.template_workbook_blob:
        filename = workflow.template_workbook_filename or f"visit-session-{session.id}.xlsx"
        workbook_bytes = workflow.template_workbook_blob
    else:
        sample_types = [sample_type.name for sample_type in sample_service.list_sample_types(db)]
        studies = [study.name for study in sample_service.list_studies(db)]
        boxes = [box.display_name for box in storage_service.list_boxes(db)]
        workbook = load_workbook(
            filename=BytesIO(
                bulk_import_service.sample_template_xlsx(sample_types=sample_types, studies=studies, boxes=boxes)
            )
        )
        workbook.active.title = VISIT_SAMPLE_SHEET
        filename = f"visit-session-{session.id}.xlsx"
        buffer = BytesIO()
        workbook.save(buffer)
        workbook_bytes = buffer.getvalue()

    session.generated_workbook_filename = filename
    _mark_step(session, STEP_TEMPLATE_DOWNLOADED)
    session.updated_at = datetime.utcnow()
    db.add(session)
    db.commit()
    return filename, workbook_bytes


def visit_workbook_to_payload(file_bytes: bytes, *, uploaded_filename: str | None = None) -> str:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sample_sheet = _find_sample_import_sheet(workbook)
    payload = {
        "uploaded_filename": uploaded_filename,
        "sample_headers": [
            _cell_text(sample_sheet.cell(row=1, column=index).value) or ""
            for index in range(1, len(bulk_import_service.SAMPLE_HEADERS) + 1)
        ],
        "sample_rows": _sheet_rows(sample_sheet),
    }
    return json.dumps(payload)


def preview_visit_workbook(
    db: Session,
    session_id: int,
    raw_payload: str,
    *,
    persist: bool = True,
    uploaded_filename: str | None = None,
) -> VisitWorkbookPreview:
    session = _get_session(db, session_id)
    parsed = _parse_raw_payload(raw_payload)
    sample_raw_payload = parsed.get("sample_raw_payload") or _sample_rows_to_raw_payload(
        parsed.get("sample_headers") or [],
        parsed.get("sample_rows") or [],
    )
    parsed["sample_raw_payload"] = sample_raw_payload
    sample_preview = bulk_import_service.preview_sample_import(db, sample_raw_payload)
    if persist:
        session.uploaded_workbook_filename = uploaded_filename or parsed.get("uploaded_filename")
        session.set_uploaded_workbook_payload(parsed)
        _mark_step(session, STEP_WORKBOOK_UPLOADED)
        session.updated_at = datetime.utcnow()
        db.add(session)
        db.commit()
        db.refresh(session)
    return VisitWorkbookPreview(
        session=_build_session_view(session),
        raw_payload=raw_payload,
        uploaded_filename=uploaded_filename or parsed.get("uploaded_filename"),
        sample_preview=sample_preview,
    )


def commit_visit_workbook(
    db: Session,
    session_id: int,
    raw_payload: str,
    *,
    uploaded_filename: str | None,
    user: models.User | None,
) -> VisitWorkbookCommitResult:
    preview = preview_visit_workbook(
        db,
        session_id,
        raw_payload,
        persist=False,
        uploaded_filename=uploaded_filename,
    )
    session = _get_session(db, session_id)
    if session.samples:
        raise VisitWorkflowError("Samples have already been committed for this visit session")
    if preview.sample_preview.global_errors or preview.sample_preview.invalid_rows:
        return VisitWorkbookCommitResult(
            session=_build_session_view(session),
            imported_rows=0,
            skipped_rows=0,
            failed_rows=0,
            global_errors=list(preview.sample_preview.global_errors) or ["Commit blocked until every workbook row is valid."],
        )

    parsed = _parse_raw_payload(raw_payload)
    session.uploaded_workbook_filename = uploaded_filename or parsed.get("uploaded_filename")
    session.set_uploaded_workbook_payload(parsed)
    session.updated_at = datetime.utcnow()

    if preview.sample_preview.total_rows == 0:
        session.status = models.VisitSessionStatus.completed
        session.completed_at = datetime.utcnow()
        db.add(session)
        db.commit()
        db.refresh(session)
        return VisitWorkbookCommitResult(
            session=_build_session_view(session),
            imported_rows=0,
            skipped_rows=0,
            failed_rows=0,
            global_errors=[],
        )

    sample_types = {item.name.lower(): item.id for item in sample_service.list_sample_types(db)}
    studies = {item.name.lower(): item.id for item in sample_service.list_studies(db)}
    position_lookup = _build_position_lookup(db)
    created_sample_ids: list[int] = []
    group_payload = _visit_submission_group_payload(session, preview.sample_preview.total_rows)

    try:
        for row in preview.sample_preview.rows:
            payload = SampleCreateInput(
                sample_id=row.sample_id or "",
                sample_type_id=sample_types.get((row.sample_type or "").lower()),
                study_id=studies.get((row.study or "").lower()),
                visit_label=row.visit,
                timepoint_label=row.timepoint,
                aliquot_number=_parse_int(row.aliquot),
                hemolysis_classification=_parse_float(row.hemolysis),
                study_role=row.study_role or "current",
                volume=_parse_float(row.volume),
                volume_units=row.volume_units or "mL",
                thaw_count=_parse_int(row.thaw_count) or 0,
                notes=row.notes,
                collection_at=_parse_datetime(row.collection_at),
            )
            sample = sample_service.create_sample(db, payload, user, commit=False, event_payload=group_payload)
            if row.assigned_box_name and row.assigned_position:
                position_id = position_lookup.get((row.assigned_box_name.lower(), row.assigned_position.upper()))
                if position_id is None:
                    raise VisitWorkflowError("Assigned visit workbook position was not found during commit")
                sample_service.place_sample(
                    db,
                    sample.id,
                    PlaceSampleInput(position_id=position_id),
                    user,
                    commit=False,
                    event_payload=group_payload,
                )
            visit_note = f"Created from visit session #{session.id} for participant {session.participant_id}"
            db.add(models.SampleNoteEntry(sample_id=sample.id, user_id=user.id if user else None, text=visit_note))
            db.add(models.VisitSessionSample(visit_session_id=session.id, sample_id=sample.id))
            created_sample_ids.append(sample.id)

        session.imported_rows = len(created_sample_ids)
        session.skipped_rows = 0
        session.failed_rows = 0
        session.status = models.VisitSessionStatus.completed
        session.completed_at = datetime.utcnow()
        _mark_step(session, STEP_WORKBOOK_UPLOADED)
        _mark_step(session, STEP_SAMPLE_IMPORT_COMMITTED)
        db.add(session)
        db.commit()
        db.refresh(session)
    except Exception:
        db.rollback()
        raise

    return VisitWorkbookCommitResult(
        session=_build_session_view(_get_session(db, session_id)),
        imported_rows=len(created_sample_ids),
        skipped_rows=0,
        failed_rows=0,
        global_errors=[],
    )


def complete_visit_session(
    db: Session,
    session_id: int,
    data: VisitSessionCompleteInput,
) -> VisitSessionView:
    session = _get_session(db, session_id)
    session.status = models.VisitSessionStatus.completed
    session.completion_note = _clean_text(data.completion_note)
    session.completed_at = datetime.utcnow()
    session.updated_at = datetime.utcnow()
    db.add(session)
    db.commit()
    db.refresh(session)
    return _build_session_view(session)


def _build_quick_links(quick_links) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    for index, item in enumerate(quick_links or [], start=1):
        label = _clean_text(getattr(item, "label", None))
        url = _clean_text(getattr(item, "url", None))
        if not label and not url:
            continue
        if not label or not url:
            raise VisitWorkflowError("Each quick link must include both a title and a URL")
        links.append({"key": f"link_{index}", "label": label, "url": url})
    return links


def _get_or_create_workflow(db: Session, study_id: int) -> models.StudyWorkflow:
    study = db.get(models.Study, study_id)
    if study is None:
        raise VisitWorkflowError("Study was not found")
    workflow = db.execute(
        select(models.StudyWorkflow).where(models.StudyWorkflow.study_id == study_id).options(joinedload(models.StudyWorkflow.study))
    ).scalar_one_or_none()
    if workflow is None:
        workflow = models.StudyWorkflow(study_id=study_id, label=f"{study.name} Workflow", is_active=False)
        workflow.set_quick_links([])
        workflow.set_sample_template_config({})
        workflow.set_summary_sections([])
        db.add(workflow)
        db.commit()
        db.refresh(workflow)
    return workflow


def _get_active_workflow_for_study(db: Session, study_id: int) -> models.StudyWorkflow:
    workflow = _get_or_create_workflow(db, study_id)
    if not workflow.is_active:
        raise VisitWorkflowError("This study workflow is not active")
    return _hydrate_workflow(db, workflow.id)


def _list_workflows(db: Session, *, active_only: bool) -> list[models.StudyWorkflow]:
    stmt = select(models.StudyWorkflow).options(joinedload(models.StudyWorkflow.study)).order_by(models.StudyWorkflow.label.asc())
    if active_only:
        stmt = stmt.where(models.StudyWorkflow.is_active.is_(True))
    return list(db.execute(stmt).unique().scalars().all())


def _hydrate_workflow(db: Session, workflow_id: int) -> models.StudyWorkflow:
    workflow = db.execute(
        select(models.StudyWorkflow)
        .where(models.StudyWorkflow.id == workflow_id)
        .options(joinedload(models.StudyWorkflow.study))
    ).scalar_one_or_none()
    if workflow is None:
        raise VisitWorkflowError("Study workflow was not found")
    return workflow


def _get_session(db: Session, session_id: int) -> models.VisitSession:
    stmt = (
        select(models.VisitSession)
        .where(models.VisitSession.id == session_id)
        .options(
            joinedload(models.VisitSession.study),
            joinedload(models.VisitSession.workflow).joinedload(models.StudyWorkflow.study),
            joinedload(models.VisitSession.operator),
            selectinload(models.VisitSession.samples)
            .joinedload(models.VisitSessionSample.sample)
            .joinedload(models.Sample.sample_type),
            selectinload(models.VisitSession.samples)
            .joinedload(models.VisitSessionSample.sample)
            .selectinload(models.Sample.location)
            .joinedload(models.SampleLocation.position)
            .joinedload(models.StoragePosition.box),
        )
    )
    session = db.execute(stmt).unique().scalar_one_or_none()
    if session is None:
        raise VisitWorkflowError("Visit session was not found")
    return session


def _build_workflow_view(workflow: models.StudyWorkflow) -> StudyWorkflowView:
    return StudyWorkflowView(
        id=workflow.id,
        study_id=workflow.study_id,
        study_name=workflow.study.display_name,
        label=workflow.label,
        description=workflow.description,
        is_active=workflow.is_active,
        quick_links=[StudyWorkflowQuickLink.model_validate(link) for link in workflow.quick_links],
        sample_template_config=workflow.sample_template_config or {},
        summary_sections=[],
        template_workbook_filename=workflow.template_workbook_filename,
        has_template_workbook=workflow.has_template_workbook,
        created_at=workflow.created_at,
        updated_at=workflow.updated_at,
    )


def _build_session_view(session: models.VisitSession) -> VisitSessionView:
    created_samples = []
    for item in session.samples:
        sample = item.sample
        created_samples.append(
            VisitSessionSampleLink(
                id=item.id,
                sample_id=item.sample_id,
                created_at=item.created_at,
                sample_identifier=sample.sample_id if sample else None,
                sample_type_name=sample.sample_type.name if sample and sample.sample_type else None,
                location_label=sample.location.position.label if sample and sample.location else None,
                detail_url=f"/samples/{sample.id}" if sample else None,
            )
        )
    return VisitSessionView(
        id=session.id,
        study_id=session.study_id,
        study_name=session.study.display_name,
        workflow_id=session.workflow_id,
        workflow_label=session.workflow.label,
        workflow_description=session.workflow.description,
        participant_id=session.participant_id,
        visit_date=session.visit_date,
        operator_user_id=session.operator_user_id,
        operator_name=(session.operator.full_name or session.operator.username) if session.operator else None,
        status=session.status.value,
        session_notes=session.session_notes,
        deviation_notes=None,
        completion_note=None,
        generated_workbook_filename=session.generated_workbook_filename,
        uploaded_workbook_filename=session.uploaded_workbook_filename,
        step_status={key: str(value) for key, value in (session.step_status or {}).items()},
        imported_rows=session.imported_rows,
        skipped_rows=session.skipped_rows,
        failed_rows=session.failed_rows,
        completed_at=session.completed_at,
        created_at=session.created_at,
        updated_at=session.updated_at,
        quick_links=[StudyWorkflowQuickLink.model_validate(link) for link in session.workflow.quick_links],
        summary_sections=[],
        created_samples=created_samples,
    )


def _visit_submission_group_payload(session: models.VisitSession, sample_count: int) -> dict[str, str | int | None]:
    title = f"{session.study.display_name} visit submission"
    return {
        "batch_group_kind": "visit_workflow",
        "batch_group_id": str(session.id),
        "batch_group_title": title,
        "batch_action_label": "Visit submission",
        "batch_workflow_label": "Visit workflow",
        "batch_sample_count": sample_count,
        "batch_count_label": "Created samples",
        "visit_session_id": session.id,
        "workflow_label": session.workflow.label,
        "study": session.study.display_name,
        "participant_id": session.participant_id,
        "visit_date": session.visit_date.isoformat() if session.visit_date else None,
        "uploaded_workbook_filename": session.uploaded_workbook_filename,
        "session_notes": session.session_notes,
    }


def _mark_step(session: models.VisitSession, step_key: str) -> None:
    payload = session.step_status or {}
    payload[step_key] = datetime.utcnow().isoformat()
    session.set_step_status(payload)


def _parse_raw_payload(raw_payload: str) -> dict:
    try:
        payload = json.loads(raw_payload or "{}")
    except json.JSONDecodeError as exc:
        raise VisitWorkflowError("Uploaded visit workbook payload is invalid") from exc
    if not isinstance(payload, dict):
        raise VisitWorkflowError("Uploaded visit workbook payload is invalid")
    return payload


def _sheet_rows(sheet) -> list[dict[str, str | None]]:
    headers = [_cell_text(sheet.cell(row=1, column=index).value) or "" for index in range(1, len(bulk_import_service.SAMPLE_HEADERS) + 1)]
    rows: list[dict[str, str | None]] = []
    for row_index in range(2, sheet.max_row + 1):
        values = {
            header: _cell_text(sheet.cell(row=row_index, column=index).value)
            for index, header in enumerate(headers, start=1)
            if header
        }
        if not any(values.values()):
            continue
        rows.append(values)
    return rows


def _sample_rows_to_raw_payload(
    headers: list[str],
    rows: list[dict[str, str | None]],
) -> str:
    active_headers = [header for header in headers if header]
    if not active_headers:
        active_headers = list(bulk_import_service.SAMPLE_HEADERS)
    lines = [",".join(active_headers)]
    for row in rows:
        if not _row_has_meaningful_input(row):
            continue
        line = ",".join(_csv_escape(_clean_text(row.get(header))) for header in active_headers)
        lines.append(line)
    return "\n".join(lines)


def _row_has_meaningful_input(row: dict[str, str | None]) -> bool:
    for value in row.values():
        if _clean_text(value):
            return True
    return False


def _validate_visit_template_workbook(file_bytes: bytes) -> bytes:
    workbook = load_workbook(filename=BytesIO(file_bytes))
    sample_sheet = _find_sample_import_sheet(workbook)
    headers = [_cell_text(sample_sheet.cell(row=1, column=index).value) or "" for index in range(1, len(bulk_import_service.SAMPLE_HEADERS) + 1)]
    if headers != list(bulk_import_service.SAMPLE_HEADERS):
        raise VisitWorkflowError(
            "The uploaded template must include a sample_Import sheet whose first row matches the bulk sample upload template."
        )
    return file_bytes


def _find_sample_import_sheet(workbook):
    exact = next((name for name in workbook.sheetnames if name == VISIT_SAMPLE_SHEET), None)
    if exact:
        return workbook[exact]
    fallback = next((name for name in workbook.sheetnames if name.strip().lower() == VISIT_SAMPLE_SHEET.lower()), None)
    if fallback:
        return workbook[fallback]
    raise VisitWorkflowError("Visit workbook must include a sample_Import sheet.")


def _csv_escape(value: str | None) -> str:
    text = value or ""
    if any(char in text for char in [",", '"', "\n"]):
        return '"' + text.replace('"', '""') + '"'
    return text


def _cell_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%y %H:%M")
    text = str(value).strip()
    return text or None


def _build_position_lookup(db: Session) -> dict[tuple[str, str], int]:
    lookup: dict[tuple[str, str], int] = {}
    for box in storage_service.list_boxes(db):
        view = storage_service.get_box_view(db, box.id)
        for position in view.positions:
            lookup[(view.box_name.lower(), position.label.upper())] = position.id
    return lookup


def _parse_int(value: str | None) -> int | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _parse_float(value: str | None) -> float | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_datetime(value: str | None) -> datetime | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        return datetime.strptime(text, "%m/%d/%y %H:%M")
    except ValueError:
        return None


def _clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
