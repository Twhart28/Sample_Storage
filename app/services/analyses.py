from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.domain import models
from app.repositories import samples as sample_repository
from app.schemas import (
    AnalysisBatchCreateInput,
    AnalysisBatchItemInput,
    AnalysisBatchResult,
    AnalysisImportCommitInput,
    AnalysisImportCommitResult,
    AnalysisImportPreview,
    AnalysisImportRow,
    AnalysisPreviewRequest,
    AnalysisPreviewResponse,
    AnalysisPreviewSample,
)
from app.services import samples as sample_service
from app.services import storage as storage_service

DATE_FORMAT = "%m/%d/%y %H:%M"
WORKBOOK_SHEET_NAME = "Analysis Log"
LOOKUP_SHEET_NAME = "_lists"
TABLE_HEADER_ROW = 6
DATA_START_ROW = 7
ENTRY_ROWS = 500
HEADERS = [
    "sample_pk",
    "sample_id",
    "sample_type",
    "current_box",
    "current_position",
    "current_volume",
    "volume_units",
    "remaining_volume",
    "returned_to_storage",
    "return_box",
    "return_position",
    "thaw_increment",
    "sample_notes",
]
REQUIRED_HEADERS = {"sample_pk", "remaining_volume"}
HEADER_COMMENT_SIZES = {
    "sample_pk": (230, 70),
    "sample_id": (220, 70),
    "sample_type": (220, 70),
    "current_box": (360, 85),
    "current_position": (230, 70),
    "current_volume": (240, 70),
    "volume_units": (220, 70),
    "remaining_volume": (290, 85),
    "returned_to_storage": (320, 85),
    "return_box": (340, 85),
    "return_position": (300, 85),
    "thaw_increment": (260, 85),
    "sample_notes": (300, 85),
}
HEADER_COMMENTS = {
    "sample_pk": "Required. Internal sample key used to match this row back to the correct sample on upload.",
    "sample_id": "Read-only identity reference for staff review. Changing this will cause a validation error.",
    "sample_type": "Read-only sample type reference for staff review.",
    "current_box": "Read-only current box name for stale-workbook checks.",
    "current_position": "Read-only current position label for stale-workbook checks.",
    "current_volume": "Read-only current remaining volume at log generation time.",
    "volume_units": "Read-only volume units for this sample.",
    "remaining_volume": "Required. Enter the volume left after analysis. This updates the sample volume but does not archive it.",
    "returned_to_storage": "Optional. Use yes or no. Blank defaults to yes.",
    "return_box": "Optional. Leave blank to return to the original box. If you enter a box, you must also enter a return position.",
    "return_position": "Optional. Leave blank to return to the original position. If return_box is blank and you enter a position, the current box is used.",
    "thaw_increment": "Optional. Whole number. Blank defaults to 1.",
    "sample_notes": "Optional. Stored on the sample note log and analysis event.",
}
TEMPLATE_WIDTHS = {
    "A": 10,
    "B": 16,
    "C": 16,
    "D": 24,
    "E": 14,
    "F": 16,
    "G": 12,
    "H": 18,
    "I": 18,
    "J": 20,
    "K": 16,
    "L": 14,
    "M": 34,
}


class AnalysisError(Exception):
    pass


@dataclass
class PositionPlan:
    box_id: int
    box_name: str
    box_lookup_names: set[str]
    positions_by_label: dict[str, models.StoragePosition]


@dataclass
class PlannedItem:
    item: AnalysisBatchItemInput
    sample: models.Sample
    from_position: models.StoragePosition
    to_position: models.StoragePosition | None
    remaining_volume: float | None
    returned_to_storage: bool
    thaw_increment: int
    sample_notes: str | None


def preview_samples(db: Session, payload: AnalysisPreviewRequest) -> AnalysisPreviewResponse:
    sample_ids = _unique_sample_ids(payload.sample_ids)
    samples = sample_repository.get_by_ids(db, sample_ids)
    sample_lookup = {sample.id: sample for sample in samples}
    return AnalysisPreviewResponse(
        samples=[
            _preview_sample(sample_lookup[sample_id]) if sample_id in sample_lookup else _missing_preview_sample(sample_id)
            for sample_id in sample_ids
        ]
    )


def generate_analysis_log_xlsx(db: Session, sample_ids: list[int]) -> bytes:
    samples = sample_repository.get_by_ids(db, _unique_sample_ids(sample_ids))
    eligible = [sample for sample in samples if _is_eligible(sample)]
    if not eligible:
        raise AnalysisError("Select at least one placed, active sample before generating a log")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKBOOK_SHEET_NAME
    sheet.freeze_panes = f"A{DATA_START_ROW}"

    header_fill = PatternFill(fill_type="solid", fgColor="1F5C4B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    notes_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sheet["A1"] = "Batch header (optional)"
    sheet["A2"] = "analysis_type"
    sheet["A3"] = "performed_at"
    sheet["A4"] = "overall_notes"
    sheet["A1"].font = Font(bold=True)
    for label_cell in ("A2", "A3", "A4"):
        sheet[label_cell].font = Font(bold=True)
    sheet["B4"].alignment = notes_alignment

    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=TABLE_HEADER_ROW, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.comment = Comment(HEADER_COMMENTS.get(header, ""), "Sample Storage")
        cell.comment.width, cell.comment.height = HEADER_COMMENT_SIZES.get(header, (260, 80))
        sheet.column_dimensions[cell.column_letter].width = TEMPLATE_WIDTHS.get(cell.column_letter, 16)

    lookup_sheet = workbook.create_sheet(LOOKUP_SHEET_NAME)
    lookup_sheet.sheet_state = "hidden"
    lookup_sheet["A1"] = "returned_to_storage"
    lookup_sheet["A2"] = "yes"
    lookup_sheet["A3"] = "no"
    lookup_sheet["B1"] = "boxes"
    box_names = [box.display_name for box in storage_service.list_boxes(db)]
    for index, box_name in enumerate(box_names, start=2):
        lookup_sheet.cell(row=index, column=2, value=box_name)

    returned_validation = DataValidation(type="list", formula1=f"'{LOOKUP_SHEET_NAME}'!$A$2:$A$3", allow_blank=True)
    returned_validation.error = "Use yes or no."
    sheet.add_data_validation(returned_validation)
    returned_validation.add(f"I{DATA_START_ROW}:I{ENTRY_ROWS}")

    thaw_validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    thaw_validation.error = "Use a whole number greater than or equal to 0."
    sheet.add_data_validation(thaw_validation)
    thaw_validation.add(f"L{DATA_START_ROW}:L{ENTRY_ROWS}")

    if box_names:
        box_formula = f"'{LOOKUP_SHEET_NAME}'!$B$2:$B${len(box_names) + 1}"
        box_validation = DataValidation(type="list", formula1=box_formula, allow_blank=True)
        box_validation.error = "Choose a configured box."
        sheet.add_data_validation(box_validation)
        box_validation.add(f"J{DATA_START_ROW}:J{ENTRY_ROWS}")

    for row_index, sample in enumerate(eligible, start=DATA_START_ROW):
        position = sample.location.position
        current_box = position.box.display_name
        current_position = position.label
        current_volume = "" if sample.volume is None else sample.volume
        entries = [
            sample.id,
            sample.sample_id,
            sample.sample_type.name if sample.sample_type else "",
            current_box,
            current_position,
            current_volume,
            sample.volume_units or "mL",
            current_volume,
            "yes",
            "",
            "",
            1,
            "",
        ]
        for column_index, value in enumerate(entries, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if column_index in {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}:
                cell.alignment = centered_alignment
            else:
                cell.alignment = notes_alignment

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def analysis_workbook_to_payload(file_bytes: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook[WORKBOOK_SHEET_NAME] if WORKBOOK_SHEET_NAME in workbook.sheetnames else workbook.active
    headers = [_clean_text(sheet.cell(row=TABLE_HEADER_ROW, column=index).value) or "" for index in range(1, len(HEADERS) + 1)]
    rows: list[dict[str, str | int | None]] = []
    for row_number in range(DATA_START_ROW, sheet.max_row + 1):
        values = {
            header: _xlsx_value_to_text(sheet.cell(row=row_number, column=index).value)
            for index, header in enumerate(headers, start=1)
            if header
        }
        if not any(_clean_text(value) for value in values.values()):
            continue
        values["row_number"] = row_number
        rows.append(values)
    payload = {
        "analysis_type": _clean_text(sheet["B2"].value),
        "performed_at": _xlsx_value_to_text(sheet["B3"].value),
        "overall_notes": _clean_text(sheet["B4"].value),
        "headers": headers,
        "rows": rows,
    }
    return json.dumps(payload)


def preview_analysis_import(db: Session, raw_payload: str) -> AnalysisImportPreview:
    parsed = _parse_import_payload(raw_payload)
    preview = AnalysisImportPreview(
        raw_payload=raw_payload,
        analysis_type=parsed.get("analysis_type"),
        performed_at=parsed.get("performed_at"),
        overall_notes=parsed.get("overall_notes"),
        headers=parsed.get("headers") or [],
        global_errors=[],
    )

    missing_headers = sorted(REQUIRED_HEADERS.difference(preview.headers))
    unknown_headers = sorted(set(preview.headers).difference(HEADERS))
    if missing_headers:
        preview.global_errors.append(f"Missing required columns: {', '.join(missing_headers)}")
    if unknown_headers:
        preview.global_errors.append(f"Unknown columns: {', '.join(unknown_headers)}")
    if preview.performed_at and _parse_datetime(preview.performed_at) is None:
        preview.global_errors.append(f"Performed at must use {DATE_FORMAT}")

    raw_rows = parsed.get("rows") or []
    preview.rows = [
        AnalysisImportRow(
            row_number=_coerce_row_number(row.get("row_number"), index),
            sample_pk=_clean_text(row.get("sample_pk")),
            sample_id=_clean_text(row.get("sample_id")),
            sample_type=_clean_text(row.get("sample_type")),
            current_box=_clean_text(row.get("current_box")),
            current_position=_clean_text(row.get("current_position")),
            current_volume=_clean_text(row.get("current_volume")),
            volume_units=_clean_text(row.get("volume_units")),
            remaining_volume=_clean_text(row.get("remaining_volume")),
            returned_to_storage=_clean_text(row.get("returned_to_storage")),
            return_box=_clean_text(row.get("return_box")),
            return_position=_clean_text(row.get("return_position")),
            thaw_increment=_clean_text(row.get("thaw_increment")),
            sample_notes=_clean_text(row.get("sample_notes")),
        )
        for index, row in enumerate(raw_rows, start=DATA_START_ROW)
    ]

    if preview.global_errors:
        for row in preview.rows:
            row.errors.extend(preview.global_errors)
            row.status = "invalid"
            row.valid = False
        _finalize_preview_counts(preview)
        return preview

    sample_ids = [int(row.sample_pk) for row in preview.rows if (row.sample_pk or "").isdigit()]
    sample_lookup = {sample.id: sample for sample in sample_repository.get_by_ids(db, _unique_sample_ids(sample_ids))}
    duplicate_sample_ids = _duplicate_text_values([row.sample_pk for row in preview.rows])
    box_plans = _build_box_plans(db)
    reserved_targets: dict[int, int] = {}

    for row in preview.rows:
        _validate_import_row(
            row,
            sample_lookup=sample_lookup,
            duplicate_sample_ids=duplicate_sample_ids,
            box_plans=box_plans,
            reserved_targets=reserved_targets,
        )

    _finalize_preview_counts(preview)
    return preview


def commit_analysis_import(
    db: Session,
    data: AnalysisImportCommitInput,
    user: models.User | None,
) -> AnalysisImportCommitResult:
    preview = preview_analysis_import(db, data.raw_payload)
    result = AnalysisImportCommitResult(
        rows=[row.model_copy(deep=True) for row in preview.rows],
        global_errors=list(preview.global_errors),
    )
    if preview.global_errors or preview.invalid_rows or preview.valid_rows == 0:
        result.global_errors.append("Commit blocked until every analysis row is valid.")
        result.skipped_rows = len(result.rows)
        for row in result.rows:
            if row.valid:
                row.status = "skipped"
        return result

    payload = _preview_to_batch_input(db, preview)
    batch_result = submit_analysis_batch(db, payload, user)
    for row in result.rows:
        row.status = "imported"
        row.valid = True
    result.imported_rows = len(result.rows)
    result.batch_id = batch_result.id
    return result


def submit_analysis_batch(
    db: Session,
    data: AnalysisBatchCreateInput,
    user: models.User | None,
) -> AnalysisBatchResult:
    items = list(data.items or [])
    if not items:
        raise AnalysisError("Add at least one sample to the analysis batch")

    sample_ids = [item.sample_id for item in items]
    if len(sample_ids) != len(set(sample_ids)):
        raise AnalysisError("Each sample can only appear once in an analysis batch")

    sample_lookup = {sample.id: sample for sample in sample_repository.get_by_ids(db, sample_ids)}
    planned_items = _plan_batch_items(db, items, sample_lookup)
    stored_analysis_type = _clean_text(data.analysis_type) or "Analysis"
    stored_performed_at = data.performed_at or datetime.utcnow()
    batch = models.AnalysisBatch(
        analysis_type=stored_analysis_type,
        performed_at=stored_performed_at,
        overall_notes=_clean_text(data.overall_notes),
        user_id=user.id if user else None,
        created_at=datetime.utcnow(),
    )

    try:
        db.add(batch)
        db.flush()
        returned_count = sum(1 for planned in planned_items if planned.returned_to_storage)
        out_for_analysis_count = len(planned_items) - returned_count
        batch_group_title = f"{stored_analysis_type} batch #{batch.id}"

        for planned in planned_items:
            sample = planned.sample
            before_snapshot = sample_service._sample_audit_snapshot(sample)
            from_path = storage_service.storage_path_for_position(planned.from_position)
            to_path = storage_service.storage_path_for_position(planned.to_position) if planned.to_position else None
            sample.volume = planned.remaining_volume
            sample.thaw_count = max(sample.thaw_count + planned.thaw_increment, 0)
            sample.updated_at = datetime.utcnow()

            if planned.returned_to_storage and planned.to_position is not None:
                sample.is_archived = False
                sample.is_out_for_analysis = False
                if sample.location is not None:
                    sample.location.position_id = planned.to_position.id
                    sample.location.placed_at = datetime.utcnow()
                else:
                    db.add(
                        models.SampleLocation(
                            sample_id=sample.id,
                            position_id=planned.to_position.id,
                            placed_at=datetime.utcnow(),
                        )
                    )
            else:
                sample.is_archived = False
                sample.is_out_for_analysis = True
                if sample.location is not None:
                    db.delete(sample.location)
                    sample.location = None

            sample_service._validate_sample_custody(sample)
            db.add(sample)
            db.add(
                models.AnalysisItem(
                    batch_id=batch.id,
                    sample_id=sample.id,
                    from_position_id=planned.from_position.id,
                    to_position_id=planned.to_position.id if planned.to_position else None,
                    remaining_volume=planned.remaining_volume,
                    volume_units=sample.volume_units or "mL",
                    thaw_increment=planned.thaw_increment,
                    returned_to_storage=planned.returned_to_storage,
                    sample_notes=planned.sample_notes,
                    created_at=datetime.utcnow(),
                )
            )
            note_text = _build_note_text(
                analysis_type=batch.analysis_type,
                sample=sample,
                remaining_volume=planned.remaining_volume,
                returned_to_storage=planned.returned_to_storage,
                to_path=to_path,
                sample_notes=planned.sample_notes,
            )
            db.add(models.SampleNoteEntry(sample_id=sample.id, user_id=user.id if user else None, text=note_text))
            db.flush()
            db.refresh(sample)

            after_snapshot = sample_service._sample_audit_snapshot(sample)
            changes = sample_service._sample_snapshot_changes(before_snapshot, after_snapshot)
            event = models.Event(
                event_type=models.EventType.analyze_sample,
                user_id=user.id if user else None,
                sample_id=sample.id,
                from_position_id=planned.from_position.id,
                to_position_id=planned.to_position.id if planned.to_position else None,
                created_at=datetime.utcnow(),
            )
            event.set_payload(
                {
                    "analysis_batch_id": batch.id,
                    "batch_group_kind": "analysis_batch",
                    "batch_group_id": str(batch.id),
                    "batch_group_title": batch_group_title,
                    "batch_sample_count": len(planned_items),
                    "batch_returned_count": returned_count,
                    "batch_out_for_analysis_count": out_for_analysis_count,
                    "analysis_type": batch.analysis_type,
                    "performed_at": batch.performed_at.isoformat() if batch.performed_at else None,
                    "overall_notes": batch.overall_notes,
                    "sample_notes": planned.sample_notes,
                    "returned_to_storage": planned.returned_to_storage,
                    "disposition": "returned_to_storage" if planned.returned_to_storage else "out_for_analysis",
                    "from_position_id": planned.from_position.id,
                    "from_path": from_path,
                    "to_position_id": planned.to_position.id if planned.to_position else None,
                    "to_path": to_path,
                    "before": before_snapshot,
                    "after": after_snapshot,
                    "changes": changes,
                    "thaw_increment": planned.thaw_increment,
                }
            )
            db.add(event)

        db.commit()
        db.refresh(batch)
    except Exception:
        db.rollback()
        raise

    return AnalysisBatchResult(
        id=batch.id,
        analysis_type=batch.analysis_type,
        performed_at=batch.performed_at,
        overall_notes=batch.overall_notes,
        created_at=batch.created_at,
        returned_count=returned_count,
        out_for_analysis_count=out_for_analysis_count,
        processed_count=len(planned_items),
        sample_ids=[planned.sample.id for planned in planned_items],
    )


def _preview_sample(sample: models.Sample) -> AnalysisPreviewSample:
    location_path = storage_service.storage_path_for_position(sample.location.position) if sample.location else None
    location_label = sample.location.position.label if sample.location else None
    eligible = _is_eligible(sample)
    return AnalysisPreviewSample(
        id=sample.id,
        sample_id=sample.sample_id,
        study_role=sample.study_role.value,
        custody_label=sample_service.derive_custody(sample),
        usage_label=sample_service.derive_usage(sample),
        volume=sample.volume,
        volume_units=sample.volume_units,
        thaw_count=sample.thaw_count,
        sample_type_name=sample.sample_type.name if sample.sample_type else None,
        study_name=sample.study.display_name if sample.study else None,
        location_label=location_label,
        location_path=location_path,
        location_position_id=sample.location.position_id if sample.location else None,
        eligible=eligible,
        ineligibility_reason=None if eligible else _ineligibility_reason(sample),
    )


def _missing_preview_sample(sample_id: int) -> AnalysisPreviewSample:
    return AnalysisPreviewSample(
        id=sample_id,
        sample_id=str(sample_id),
        study_role="unknown",
        custody_label="unknown",
        usage_label="unknown",
        thaw_count=0,
        eligible=False,
        ineligibility_reason="Sample was not found",
    )


def _plan_batch_items(
    db: Session,
    items: list[AnalysisBatchItemInput],
    sample_lookup: dict[int, models.Sample],
) -> list[PlannedItem]:
    box_plans = _build_box_plans(db)
    reserved_positions: dict[int, int] = {}
    planned: list[PlannedItem] = []
    for item in items:
        sample = sample_lookup.get(item.sample_id)
        if sample is None:
            raise AnalysisError(f"Sample {item.sample_id} was not found")
        if not _is_eligible(sample):
            raise AnalysisError(f"Sample {sample.sample_id} must be placed and not archived before analysis")
        if item.thaw_increment < 0:
            raise AnalysisError(f"Sample {sample.sample_id} thaw increment cannot be negative")
        if item.remaining_volume is not None and item.remaining_volume < 0:
            raise AnalysisError(f"Sample {sample.sample_id} remaining volume cannot be negative")

        from_position = sample.location.position
        remaining_volume = item.remaining_volume if item.remaining_volume is not None else sample.volume
        returned_to_storage = bool(item.returned_to_storage)

        to_position: models.StoragePosition | None = None
        if returned_to_storage:
            target_id = item.return_position_id or from_position.id
            to_position = _resolve_return_position(target_id, box_plans)
            if to_position.location and to_position.location.sample_id != sample.id:
                raise AnalysisError(f"Return position {to_position.label} is already occupied")
            other_sample_id = reserved_positions.get(to_position.id)
            if other_sample_id is not None and other_sample_id != sample.id:
                raise AnalysisError("Two analysis rows cannot return to the same position")
            reserved_positions[to_position.id] = sample.id

        planned.append(
            PlannedItem(
                item=item,
                sample=sample,
                from_position=from_position,
                to_position=to_position,
                remaining_volume=remaining_volume,
                returned_to_storage=returned_to_storage,
                thaw_increment=item.thaw_increment,
                sample_notes=_clean_text(item.sample_notes),
            )
        )
    return planned


def _validate_import_row(
    row: AnalysisImportRow,
    *,
    sample_lookup: dict[int, models.Sample],
    duplicate_sample_ids: set[str],
    box_plans: dict[int, PositionPlan],
    reserved_targets: dict[int, int],
) -> None:
    sample_pk = _parse_int(row.sample_pk)
    if sample_pk is None:
        row.errors.append("sample_pk is required")
        row.status = "invalid"
        row.valid = False
        return
    if row.sample_pk in duplicate_sample_ids:
        row.errors.append("sample_pk is duplicated in this workbook")

    sample = sample_lookup.get(sample_pk)
    if sample is None:
        row.errors.append("Sample was not found")
        row.status = "invalid"
        row.valid = False
        return
    if not _is_eligible(sample):
        row.errors.append("Sample must be placed and not archived before analysis")

    expected_box = sample.location.position.box.display_name if sample.location else None
    expected_position = sample.location.position.label if sample.location else None
    expected_volume = sample.volume
    expected_units = sample.volume_units or "mL"
    expected_type = sample.sample_type.name if sample.sample_type else None

    if row.sample_id and row.sample_id != sample.sample_id:
        row.errors.append("Sample ID does not match the current record")
    if row.sample_type and row.sample_type != (expected_type or ""):
        row.errors.append("Sample type does not match the current record")
    if row.current_box and row.current_box != (expected_box or ""):
        row.errors.append("Current box does not match the current record")
    if row.current_position and row.current_position.upper() != (expected_position or "").upper():
        row.errors.append("Current position does not match the current record")
    if row.volume_units and row.volume_units != expected_units:
        row.errors.append("Volume units do not match the current record")
    if row.current_volume and not _volumes_match(row.current_volume, expected_volume):
        row.errors.append("Current volume does not match the current record")

    raw_returned = _clean_text(row.returned_to_storage)
    returned_to_storage = _parse_yes_no(raw_returned)
    if raw_returned and returned_to_storage is None:
        row.errors.append("returned_to_storage must be yes or no")
        returned_to_storage = True
    elif returned_to_storage is None:
        returned_to_storage = True
    remaining_volume = _parse_float(row.remaining_volume)
    if row.remaining_volume in (None, ""):
        remaining_volume = expected_volume
    if remaining_volume is None:
        row.errors.append("Remaining volume is required")
    elif remaining_volume < 0:
        row.errors.append("Remaining volume must be zero or greater")

    thaw_increment = _parse_int(row.thaw_increment)
    if row.thaw_increment in (None, ""):
        thaw_increment = 1
    if thaw_increment is None or thaw_increment < 0:
        row.errors.append("Thaw increment must be a whole number greater than or equal to 0")

    row.returned_to_storage = "yes" if returned_to_storage else "no"
    row.remaining_volume = "" if remaining_volume is None else f"{remaining_volume:g}"
    row.thaw_increment = "" if thaw_increment is None else str(thaw_increment)

    if returned_to_storage:
        target = _resolve_target_from_row(row, sample, box_plans)
        if target is None:
            row.errors.append("Returned samples must resolve to a valid open position")
        else:
            if target.location and target.location.sample_id != sample.id:
                row.errors.append("Return position is already occupied")
            other_sample_id = reserved_targets.get(target.id)
            if other_sample_id is not None and other_sample_id != sample.id:
                row.errors.append("Return position is duplicated in this workbook")
            reserved_targets[target.id] = sample.id
            row.assigned_box_name = target.box.display_name
            row.assigned_position = target.label
    else:
        row.assigned_box_name = None
        row.assigned_position = None

    row.valid = not row.errors
    row.status = "valid" if row.valid else "invalid"


def _resolve_target_from_row(
    row: AnalysisImportRow,
    sample: models.Sample,
    box_plans: dict[int, PositionPlan],
) -> models.StoragePosition | None:
    if sample.location is None:
        return None
    current_position = sample.location.position
    return_box = _clean_text(row.return_box)
    return_position = _clean_text(row.return_position)
    if not return_box and not return_position:
        return current_position
    if not return_box and return_position:
        plan = box_plans.get(current_position.box_id)
        position = _position_by_label(plan, return_position)
        if position is None:
            row.errors.append("Return position was not found in the current box")
        return position
    if return_box and not return_position:
        row.errors.append("Return position is required when a return box is selected")
        return None
    plan = _find_box_plan(box_plans, return_box)
    if plan is None:
        row.errors.append("Return box was not found")
        return None
    position = _position_by_label(plan, return_position)
    if position is None:
        row.errors.append("Return position was not found in the selected box")
    return position


def _position_by_label(plan: PositionPlan | None, label: str | None) -> models.StoragePosition | None:
    if plan is None or not label:
        return None
    return plan.positions_by_label.get(label.upper())


def _preview_to_batch_input(db: Session, preview: AnalysisImportPreview) -> AnalysisBatchCreateInput:
    box_plans = _build_box_plans(db)
    sample_ids = [int(row.sample_pk) for row in preview.rows if row.sample_pk]
    sample_lookup = {sample.id: sample for sample in sample_repository.get_by_ids(db, sample_ids)}
    items: list[AnalysisBatchItemInput] = []
    for row in preview.rows:
        sample_pk = int(row.sample_pk or "0")
        sample = sample_lookup.get(sample_pk)
        return_position_id = None
        if sample is not None and (row.returned_to_storage or "yes").lower() != "no":
            target = _resolve_target_from_row(row.model_copy(deep=True), sample, box_plans)
            return_position_id = target.id if target else None
        items.append(
            AnalysisBatchItemInput(
                sample_id=sample_pk,
                remaining_volume=_parse_float(row.remaining_volume),
                returned_to_storage=(row.returned_to_storage or "yes").lower() != "no",
                return_position_id=return_position_id,
                thaw_increment=_parse_int(row.thaw_increment) or 1,
                sample_notes=_clean_text(row.sample_notes),
            )
        )

    return AnalysisBatchCreateInput(
        analysis_type=_clean_text(preview.analysis_type),
        performed_at=_parse_datetime(preview.performed_at) if preview.performed_at else None,
        overall_notes=_clean_text(preview.overall_notes),
        items=items,
    )


def _build_box_plans(db: Session) -> dict[int, PositionPlan]:
    plans: dict[int, PositionPlan] = {}
    for box in storage_service.list_boxes(db):
        box_view = storage_service.get_box_view(db, box.id)
        if box_view is None:
            continue
        position_map: dict[str, models.StoragePosition] = {}
        for position in box_view.positions:
            resolved = storage_service.get_position(db, position.id)
            if resolved is not None:
                position_map[position.label.upper()] = resolved
        plans[box.id] = PositionPlan(
            box_id=box.id,
            box_name=box.display_name,
            box_lookup_names={box.display_name.lower(), box.name.lower()},
            positions_by_label=position_map,
        )
    return plans


def _resolve_return_position(target_id: int, box_plans: dict[int, PositionPlan]) -> models.StoragePosition:
    for plan in box_plans.values():
        for position in plan.positions_by_label.values():
            if position.id == target_id:
                return position
    raise AnalysisError("Return position was not found")


def _find_box_plan(box_plans: dict[int, PositionPlan], box_name: str | None) -> PositionPlan | None:
    if not box_name:
        return None
    normalized = box_name.strip().lower()
    for plan in box_plans.values():
        if normalized in plan.box_lookup_names:
            return plan
    return None


def _build_note_text(
    *,
    analysis_type: str | None,
    sample: models.Sample,
    remaining_volume: float | None,
    returned_to_storage: bool,
    to_path: str | None,
    sample_notes: str | None,
) -> str:
    summary = analysis_type or "Analysis"
    disposition = "returned to storage" if returned_to_storage else "out for analysis"
    volume_text = sample_service._volume_display(remaining_volume, sample.volume_units)
    parts = [f"Analysis: {summary}", f"Remaining volume {volume_text}", disposition]
    if returned_to_storage and to_path:
        parts.append(f"Location {to_path}")
    if sample_notes:
        parts.append(sample_notes)
    return " | ".join(parts)


def _parse_import_payload(raw_payload: str) -> dict:
    try:
        payload = json.loads(raw_payload or "{}")
    except json.JSONDecodeError as exc:
        raise AnalysisError("Analysis upload payload could not be parsed") from exc
    if not isinstance(payload, dict):
        raise AnalysisError("Analysis upload payload could not be parsed")
    return payload


def _finalize_preview_counts(preview: AnalysisImportPreview) -> None:
    preview.total_rows = len(preview.rows)
    preview.valid_rows = sum(1 for row in preview.rows if row.valid)
    preview.invalid_rows = preview.total_rows - preview.valid_rows


def _is_eligible(sample: models.Sample) -> bool:
    return not sample.is_archived and sample.location is not None


def _ineligibility_reason(sample: models.Sample) -> str:
    if sample.is_archived:
        return "Archived samples cannot be analyzed"
    if sample.is_out_for_analysis:
        return "Samples already out for analysis cannot be analyzed again until returned"
    if sample.location is None:
        return "Sample must be placed in storage before analysis"
    return "ineligible"


def _duplicate_text_values(values: list[str | None]) -> set[str]:
    counts: dict[str, int] = {}
    for value in values:
        normalized = _clean_text(value)
        if not normalized:
            continue
        counts[normalized] = counts.get(normalized, 0) + 1
    return {value for value, count in counts.items() if count > 1}


def _coerce_row_number(value, fallback: int) -> int:
    if isinstance(value, int):
        return value
    parsed = _parse_int(_clean_text(value))
    return parsed or fallback


def _unique_sample_ids(sample_ids: list[int]) -> list[int]:
    unique: list[int] = []
    for value in sample_ids:
        parsed = int(value)
        if parsed > 0 and parsed not in unique:
            unique.append(parsed)
    return unique


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_int(value: str | None) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _parse_float(value: str | None) -> float | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_yes_no(value: str | None) -> bool | None:
    cleaned = (_clean_text(value) or "").lower()
    if cleaned == "":
        return None
    if cleaned in {"yes", "y", "true"}:
        return True
    if cleaned in {"no", "n", "false"}:
        return False
    return None


def _parse_datetime(value: str | None) -> datetime | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return datetime.strptime(cleaned, DATE_FORMAT)
    except ValueError:
        try:
            return datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        except ValueError:
            return None


def _volumes_match(current_volume: str, expected_volume: float | None) -> bool:
    parsed = _parse_float(current_volume)
    if parsed is None and expected_volume is None:
        return True
    if parsed is None or expected_volume is None:
        return False
    return abs(parsed - expected_volume) < 1e-9


def _xlsx_value_to_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_text(value)
