from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.repositories import samples as sample_repository
from app.schemas import (
    BatchModifyImportCommitInput,
    BatchModifyImportCommitResult,
    BatchModifyImportPreview,
    BatchModifyImportRow,
    SampleSearchQuery,
    SampleUpdateInput,
)
from app.services import samples as sample_service

DATE_FORMAT = "%m/%d/%y %H:%M"
WORKBOOK_SHEET_NAME = "Batch Modify"
LOOKUP_SHEET_NAME = "_lists"
TABLE_HEADER_ROW = 1
DATA_START_ROW = 2
ENTRY_ROWS = 500
HEADERS = [
    "sample_pk",
    "sample_id",
    "current_sample_type",
    "current_study",
    "current_study_role",
    "current_visit",
    "current_timepoint",
    "current_aliquot",
    "current_hemolysis",
    "current_volume",
    "volume_units",
    "current_thaw_count",
    "current_notes",
    "current_collection_at",
    "sample_type",
    "study",
    "study_role",
    "visit",
    "timepoint",
    "aliquot",
    "hemolysis",
    "volume",
    "thaw_count",
    "notes",
    "collection_at",
]
REQUIRED_HEADERS = {"sample_pk", "thaw_count"}
HEADER_COMMENTS = {
    "sample_pk": "Required. Internal sample key used to match the row back to the correct sample.",
    "sample_id": "Read-only sample ID. Changing this causes validation to fail.",
    "current_sample_type": "Read-only current sample type at log generation time.",
    "current_study": "Read-only current study at log generation time.",
    "current_study_role": "Read-only current study role at log generation time.",
    "current_visit": "Read-only current visit label at log generation time.",
    "current_timepoint": "Read-only current timepoint label at log generation time.",
    "current_aliquot": "Read-only current aliquot at log generation time.",
    "current_hemolysis": "Read-only current hemolysis at log generation time.",
    "current_volume": "Read-only current volume at log generation time.",
    "volume_units": "Read-only volume units. Volume edits keep these units.",
    "current_thaw_count": "Read-only current thaw count at log generation time.",
    "current_notes": "Read-only current note summary at log generation time.",
    "current_collection_at": "Read-only current collection timestamp at log generation time.",
    "sample_type": "Editable target sample type. Blank clears the field.",
    "study": "Editable target study. Blank clears the field.",
    "study_role": "Editable target study role. Use current or retired.",
    "visit": "Editable target visit. Use numbers only. Blank clears the field.",
    "timepoint": "Editable target timepoint. Use numbers only. Blank clears the field.",
    "aliquot": "Editable target aliquot. Blank clears the field.",
    "hemolysis": "Editable target hemolysis from 0 to 6. Blank clears the field.",
    "volume": "Editable target volume. Blank clears the field.",
    "thaw_count": "Editable target thaw count. This field cannot be blank.",
    "notes": "Editable target note text. Blank clears the field.",
    "collection_at": f"Editable target collection timestamp in {DATE_FORMAT}. Blank clears the field.",
}
HEADER_COMMENT_SIZES = {header: (280, 85) for header in HEADERS}
TEMPLATE_WIDTHS = {
    "A": 10,
    "B": 16,
    "C": 18,
    "D": 18,
    "E": 12,
    "F": 14,
    "G": 12,
    "H": 12,
    "I": 12,
    "J": 10,
    "K": 12,
    "L": 24,
    "M": 18,
    "N": 18,
    "O": 18,
    "P": 12,
    "Q": 14,
    "R": 12,
    "S": 12,
    "T": 12,
    "U": 12,
    "V": 24,
    "W": 18,
}


class BatchModifyError(Exception):
    pass


@dataclass
class ModifyPlan:
    row: BatchModifyImportRow
    sample_id: int
    payload: SampleUpdateInput


def generate_modify_log_xlsx(db: Session, sample_ids: list[int]) -> bytes:
    samples = sample_repository.get_by_ids(db, _unique_sample_ids(sample_ids))
    if not samples:
        raise BatchModifyError("Select at least one sample before generating a batch modify log")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKBOOK_SHEET_NAME
    sheet.freeze_panes = f"A{DATA_START_ROW}"

    header_fill = PatternFill(fill_type="solid", fgColor="1F5C4B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    notes_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=TABLE_HEADER_ROW, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.comment = Comment(HEADER_COMMENTS.get(header, ""), "Sample Storage")
        cell.comment.width, cell.comment.height = HEADER_COMMENT_SIZES[header]
        sheet.column_dimensions[cell.column_letter].width = TEMPLATE_WIDTHS.get(cell.column_letter, 16)

    lookup_sheet = workbook.create_sheet(LOOKUP_SHEET_NAME)
    lookup_sheet.sheet_state = "hidden"
    sample_types = sample_service.list_sample_types(db)
    studies = sample_service.list_studies(db)
    lookup_sheet["A1"] = "sample_types"
    for index, sample_type in enumerate(sample_types, start=2):
        lookup_sheet.cell(row=index, column=1, value=sample_type.name)
    lookup_sheet["B1"] = "studies"
    for index, study in enumerate(studies, start=2):
        lookup_sheet.cell(row=index, column=2, value=study.display_name)
    lookup_sheet["C1"] = "study_roles"
    lookup_sheet["C2"] = "current"
    lookup_sheet["C3"] = "retired"

    if sample_types:
        formula = f"'{LOOKUP_SHEET_NAME}'!$A$2:$A${len(sample_types) + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a configured sample type."
        sheet.add_data_validation(validation)
        validation.add(f"N{DATA_START_ROW}:N{ENTRY_ROWS}")

    if studies:
        formula = f"'{LOOKUP_SHEET_NAME}'!$B$2:$B${len(studies) + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a configured study."
        sheet.add_data_validation(validation)
        validation.add(f"P{DATA_START_ROW}:P{ENTRY_ROWS}")

    study_role_validation = DataValidation(type="list", formula1=f"'{LOOKUP_SHEET_NAME}'!$C$2:$C$3", allow_blank=False)
    study_role_validation.error = "Use current or retired."
    sheet.add_data_validation(study_role_validation)
    study_role_validation.add(f"Q{DATA_START_ROW}:Q{ENTRY_ROWS}")

    hemolysis_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="6", allow_blank=True)
    hemolysis_validation.error = "Use a whole number from 0 to 6."
    sheet.add_data_validation(hemolysis_validation)
    hemolysis_validation.add(f"U{DATA_START_ROW}:U{ENTRY_ROWS}")

    thaw_validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
    thaw_validation.error = "Use a whole number greater than or equal to 0."
    sheet.add_data_validation(thaw_validation)
    thaw_validation.add(f"W{DATA_START_ROW}:W{ENTRY_ROWS}")

    for row_index, sample in enumerate(samples, start=DATA_START_ROW):
        entries = [
            sample.id,
            sample.sample_id,
            sample.sample_type.name if sample.sample_type else "",
            sample.study.display_name if sample.study else "",
            sample.study_role.value,
            sample.visit_label or "",
            sample.timepoint_label or "",
            "" if sample.aliquot_number is None else sample.aliquot_number,
            "" if sample.hemolysis_classification is None else sample.hemolysis_classification,
            "" if sample.volume is None else sample.volume,
            sample.volume_units or "mL",
            sample.thaw_count,
            sample.notes or "",
            sample.collection_at.strftime(DATE_FORMAT) if sample.collection_at else "",
            sample.sample_type.name if sample.sample_type else "",
            sample.study.display_name if sample.study else "",
            sample.study_role.value,
            sample.visit_label or "",
            sample.timepoint_label or "",
            "" if sample.aliquot_number is None else sample.aliquot_number,
            "" if sample.hemolysis_classification is None else sample.hemolysis_classification,
            "" if sample.volume is None else sample.volume,
            sample.thaw_count,
            sample.notes or "",
            sample.collection_at.strftime(DATE_FORMAT) if sample.collection_at else "",
        ]
        for column_index, value in enumerate(entries, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = notes_alignment if column_index in {13, 24} else centered_alignment

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def modify_workbook_to_payload(file_bytes: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook[WORKBOOK_SHEET_NAME] if WORKBOOK_SHEET_NAME in workbook.sheetnames else workbook.active
    headers = [_clean_text(sheet.cell(row=TABLE_HEADER_ROW, column=index).value) or "" for index in range(1, len(HEADERS) + 1)]
    rows: list[dict[str, str | int | None]] = []
    for row_number in range(DATA_START_ROW, sheet.max_row + 1):
        values = {
            header: _xlsx_value_to_text(sheet.cell(row=row_number, column=index).value)
            for index, header in enumerate(headers, start=1)
            if header
        }
        if not any(_clean_text(value) for value in values.values()):
            continue
        values["row_number"] = row_number
        rows.append(values)
    return json.dumps({"headers": headers, "rows": rows})


def preview_modify_import(db: Session, raw_payload: str) -> BatchModifyImportPreview:
    parsed = _parse_import_payload(raw_payload)
    preview = BatchModifyImportPreview(
        raw_payload=raw_payload,
        headers=parsed.get("headers") or [],
        global_errors=[],
    )
    missing_headers = sorted(REQUIRED_HEADERS.difference(preview.headers))
    unknown_headers = sorted(set(preview.headers).difference(HEADERS))
    if missing_headers:
        preview.global_errors.append(f"Missing required columns: {', '.join(missing_headers)}")
    if unknown_headers:
        preview.global_errors.append(f"Unknown columns: {', '.join(unknown_headers)}")

    raw_rows = parsed.get("rows") or []
    preview.rows = [
        BatchModifyImportRow(
            row_number=_coerce_row_number(row.get("row_number"), index),
            sample_pk=_clean_text(row.get("sample_pk")),
            sample_id=_clean_text(row.get("sample_id")),
            current_sample_type=_clean_text(row.get("current_sample_type")),
            current_study=_clean_text(row.get("current_study")),
            current_study_role=_clean_text(row.get("current_study_role")),
            current_visit=_clean_text(row.get("current_visit")),
            current_timepoint=_clean_text(row.get("current_timepoint")),
            current_aliquot=_clean_text(row.get("current_aliquot")),
            current_hemolysis=_clean_text(row.get("current_hemolysis")),
            current_volume=_clean_text(row.get("current_volume")),
            volume_units=_clean_text(row.get("volume_units")),
            current_thaw_count=_clean_text(row.get("current_thaw_count")),
            current_notes=_clean_text(row.get("current_notes")),
            current_collection_at=_clean_text(row.get("current_collection_at")),
            sample_type=_clean_text(row.get("sample_type")),
            study=_clean_text(row.get("study")),
            study_role=_clean_text(row.get("study_role")),
            visit=_clean_text(row.get("visit")),
            timepoint=_clean_text(row.get("timepoint")),
            aliquot=_clean_text(row.get("aliquot")),
            hemolysis=_clean_text(row.get("hemolysis")),
            volume=_clean_text(row.get("volume")),
            thaw_count=_clean_text(row.get("thaw_count")),
            notes=_clean_text(row.get("notes")),
            collection_at=_clean_text(row.get("collection_at")),
        )
        for index, row in enumerate(raw_rows, start=DATA_START_ROW)
    ]

    if preview.global_errors:
        for row in preview.rows:
            row.errors.extend(preview.global_errors)
            row.status = "invalid"
            row.valid = False
        _finalize_preview_counts(preview)
        return preview

    sample_ids = [int(row.sample_pk) for row in preview.rows if (row.sample_pk or "").isdigit()]
    sample_lookup = {sample.id: sample for sample in sample_repository.get_by_ids(db, _unique_sample_ids(sample_ids))}
    duplicate_sample_ids = _duplicate_text_values([row.sample_pk for row in preview.rows])
    sample_types = sample_service.list_sample_types(db)
    studies = sample_service.list_studies(db)
    sample_type_lookup = {item.name.strip().lower(): item.id for item in sample_types}
    study_lookup = {}
    for item in studies:
        if item.display_name:
            study_lookup[item.display_name.strip().lower()] = item.id
        if item.name:
            study_lookup[item.name.strip().lower()] = item.id
    study_role_lookup = {"current": "current", "retired": "retired"}

    target_identities: dict[int, tuple[str, int | None, str | None, str | None, int | None]] = {}
    for row in preview.rows:
        _validate_row(
            row,
            sample_lookup=sample_lookup,
            duplicate_sample_ids=duplicate_sample_ids,
            sample_type_lookup=sample_type_lookup,
            study_lookup=study_lookup,
            study_role_lookup=study_role_lookup,
            target_identities=target_identities,
        )

    _validate_identity_collisions(db, preview.rows, sample_lookup, target_identities)
    _finalize_preview_counts(preview)
    return preview


def commit_modify_import(
    db: Session,
    data: BatchModifyImportCommitInput,
    user,
) -> BatchModifyImportCommitResult:
    preview = preview_modify_import(db, data.raw_payload)
    result = BatchModifyImportCommitResult(
        rows=[row.model_copy(deep=True) for row in preview.rows],
        global_errors=list(preview.global_errors),
    )
    if preview.global_errors or preview.invalid_rows or preview.valid_rows == 0:
        result.global_errors.append("Commit blocked until every batch modify row is valid.")
        result.skipped_rows = len(result.rows)
        for row in result.rows:
            if row.valid:
                row.status = "skipped"
        return result

    plans = _preview_to_modify_plans(db, preview)
    batch_group_id = uuid4().hex
    batch_group_title = "Batch modify"
    event_payload = {
        "batch_group_kind": "batch_modify",
        "batch_group_id": batch_group_id,
        "batch_group_title": batch_group_title,
        "batch_action_label": "Batch Modify",
        "batch_workflow_label": "Batch Modify",
        "batch_sample_count": len(plans),
    }

    try:
        for plan in plans:
            sample_service.update_sample(
                db,
                plan.sample_id,
                plan.payload,
                user,
                commit=False,
                event_payload=event_payload,
            )
        db.commit()
    except Exception:
        db.rollback()
        raise

    for row in result.rows:
        row.status = "imported"
        row.valid = True
    result.imported_rows = len(result.rows)
    result.batch_group_id = batch_group_id
    return result


def _validate_row(
    row: BatchModifyImportRow,
    *,
    sample_lookup,
    duplicate_sample_ids: set[str],
    sample_type_lookup: dict[str, int],
    study_lookup: dict[str, int],
    study_role_lookup: dict[str, str],
    target_identities: dict[int, tuple[str, int | None, str | None, str | None, int | None]],
) -> None:
    sample_pk = _parse_int(row.sample_pk)
    if sample_pk is None:
        row.errors.append("sample_pk is required")
        row.status = "invalid"
        row.valid = False
        return
    if row.sample_pk in duplicate_sample_ids:
        row.errors.append("sample_pk is duplicated in this workbook")

    sample = sample_lookup.get(sample_pk)
    if sample is None:
        row.errors.append("Sample was not found")
        row.status = "invalid"
        row.valid = False
        return

    if row.sample_id and row.sample_id != sample.sample_id:
        row.errors.append("Sample ID does not match the current record")
    if row.current_sample_type and row.current_sample_type != (sample.sample_type.name if sample.sample_type else ""):
        row.errors.append("Current sample type does not match the current record")
    if row.current_study and row.current_study != (sample.study.display_name if sample.study else ""):
        row.errors.append("Current study does not match the current record")
    if row.current_study_role and row.current_study_role != sample.study_role.value:
        row.errors.append("Current study role does not match the current record")
    if row.current_visit and row.current_visit != (sample.visit_label or ""):
        row.errors.append("Current visit does not match the current record")
    if row.current_timepoint and row.current_timepoint != (sample.timepoint_label or ""):
        row.errors.append("Current timepoint does not match the current record")
    if row.current_aliquot and row.current_aliquot != ("" if sample.aliquot_number is None else str(sample.aliquot_number)):
        row.errors.append("Current aliquot does not match the current record")
    if row.current_hemolysis and row.current_hemolysis != ("" if sample.hemolysis_classification is None else str(sample.hemolysis_classification)):
        row.errors.append("Current hemolysis does not match the current record")
    if row.current_volume and not _volumes_match(row.current_volume, sample.volume):
        row.errors.append("Current volume does not match the current record")
    if row.volume_units and row.volume_units != (sample.volume_units or "mL"):
        row.errors.append("Volume units do not match the current record")
    if row.current_thaw_count and row.current_thaw_count != str(sample.thaw_count):
        row.errors.append("Current thaw count does not match the current record")
    if row.current_notes != _display_optional(sample.notes):
        row.errors.append("Current notes do not match the current record")
    if row.current_collection_at != _display_collection(sample.collection_at):
        row.errors.append("Current collection timestamp does not match the current record")

    sample_type_id = _parse_lookup_value(row.sample_type, sample_type_lookup, "Sample type", row.errors)
    study_id = _parse_lookup_value(row.study, study_lookup, "Study", row.errors)
    study_role = _parse_study_role(row.study_role, study_role_lookup, row.errors)

    visit_label = _normalize_optional_numeric(row.visit, "Visit", row.errors)
    timepoint_label = _normalize_optional_numeric(row.timepoint, "Timepoint", row.errors)
    aliquot_number = _parse_optional_int(row.aliquot, "Aliquot", row.errors)
    hemolysis = _parse_optional_hemolysis(row.hemolysis, row.errors)
    volume = _parse_optional_volume(row.volume, row.errors)
    thaw_count = _parse_required_int(row.thaw_count, "Thaw count", row.errors)
    notes = _clean_text(row.notes)
    collection_at = _parse_optional_datetime(row.collection_at, "Collection date", row.errors)

    if not row.errors:
        row.visit = visit_label or None
        row.timepoint = timepoint_label or None
        row.study_role = study_role
        row.aliquot = None if aliquot_number is None else str(aliquot_number)
        row.hemolysis = None if hemolysis is None else str(hemolysis)
        row.volume = None if volume is None else f"{volume:g}"
        row.thaw_count = str(thaw_count)
        row.notes = notes
        row.collection_at = collection_at.strftime(DATE_FORMAT) if collection_at else None
        target_identities[row.row_number] = (
            sample.sample_id,
            sample_type_id,
            visit_label,
            timepoint_label,
            aliquot_number,
        )

    row.valid = not row.errors
    row.status = "valid" if row.valid else "invalid"


def _validate_identity_collisions(db: Session, rows, sample_lookup, target_identities) -> None:
    if not target_identities:
        return
    batch_sample_ids = set(sample_lookup.keys())
    seen_targets: dict[tuple[str, int | None, str | None, str | None, int | None], int] = {}
    for row in rows:
        identity = target_identities.get(row.row_number)
        if identity is None:
            continue
        other_row_number = seen_targets.get(identity)
        if other_row_number is not None:
            row.errors.append("This ID, type, visit, timepoint, and aliquot combination is duplicated in this batch modify upload")
            other_row = next((item for item in rows if item.row_number == other_row_number), None)
            if other_row is not None:
                other_row.errors.append("This ID, type, visit, timepoint, and aliquot combination is duplicated in this batch modify upload")
        else:
            seen_targets[identity] = row.row_number

    other_samples = sample_repository.search(db, SampleSearchQuery())
    other_identities = {
        (
            sample.sample_id,
            sample.sample_type_id,
            sample.visit_label,
            sample.timepoint_label,
            sample.aliquot_number,
        ): sample.id
        for sample in other_samples
        if sample.id not in batch_sample_ids
    }
    for row in rows:
        identity = target_identities.get(row.row_number)
        if identity is not None and identity in other_identities:
            row.errors.append("This ID, type, visit, timepoint, and aliquot combination already exists")
        row.valid = not row.errors
        row.status = "valid" if row.valid else "invalid"


def _preview_to_modify_plans(db: Session, preview: BatchModifyImportPreview) -> list[ModifyPlan]:
    sample_types = sample_service.list_sample_types(db)
    studies = sample_service.list_studies(db)
    sample_type_lookup = {item.name.strip().lower(): item.id for item in sample_types}
    study_lookup = {}
    for item in studies:
        if item.display_name:
            study_lookup[item.display_name.strip().lower()] = item.id
        if item.name:
            study_lookup[item.name.strip().lower()] = item.id
    study_role_lookup = {"current": "current", "retired": "retired"}
    plans: list[ModifyPlan] = []
    for row in preview.rows:
        sample_id = int(row.sample_pk or "0")
        plans.append(
            ModifyPlan(
                row=row,
                sample_id=sample_id,
                payload=SampleUpdateInput(
                    sample_type_id=_parse_lookup_value(row.sample_type, sample_type_lookup, "Sample type", []),
                    study_id=_parse_lookup_value(row.study, study_lookup, "Study", []),
                    study_role=_parse_study_role(row.study_role, study_role_lookup, []),
                    visit_label=_clean_text(row.visit),
                    timepoint_label=_clean_text(row.timepoint),
                    aliquot_number=_parse_int(row.aliquot),
                    hemolysis_classification=_parse_int(row.hemolysis),
                    volume=_parse_float(row.volume),
                    thaw_count=_parse_int(row.thaw_count),
                    notes=_clean_text(row.notes),
                    collection_at=_parse_datetime(row.collection_at),
                ),
            )
        )
    return plans


def _parse_import_payload(raw_payload: str) -> dict:
    try:
        payload = json.loads(raw_payload or "{}")
    except json.JSONDecodeError as exc:
        raise BatchModifyError("Batch modify upload payload could not be parsed") from exc
    if not isinstance(payload, dict):
        raise BatchModifyError("Batch modify upload payload could not be parsed")
    return payload


def _finalize_preview_counts(preview: BatchModifyImportPreview) -> None:
    preview.total_rows = len(preview.rows)
    preview.valid_rows = sum(1 for row in preview.rows if row.valid)
    preview.invalid_rows = preview.total_rows - preview.valid_rows


def _parse_lookup_value(value: str | None, lookup: dict[str, int], label: str, errors: list[str]) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    match = lookup.get(cleaned.lower())
    if match is None:
        errors.append(f"{label} was not found")
    return match


def _parse_study_role(value: str | None, lookup: dict[str, str], errors: list[str]) -> str | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        errors.append("Study role is required")
        return None
    match = lookup.get(cleaned.lower())
    if match is None:
        errors.append("Study role must be current or retired")
    return match


def _normalize_optional_numeric(value: str | None, label: str, errors: list[str]) -> str | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return sample_service._normalize_numeric_label(cleaned, label)
    except sample_service.SampleError as exc:
        errors.append(str(exc))
        return None


def _parse_optional_int(value: str | None, label: str, errors: list[str]) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    parsed = _parse_int(cleaned)
    if parsed is None:
        errors.append(f"{label} must be a whole number")
    return parsed


def _parse_required_int(value: str | None, label: str, errors: list[str]) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        errors.append(f"{label} is required")
        return None
    parsed = _parse_int(cleaned)
    if parsed is None or parsed < 0:
        errors.append(f"{label} must be a whole number greater than or equal to 0")
        return None
    return parsed


def _parse_optional_hemolysis(value: str | None, errors: list[str]) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    parsed = _parse_int(cleaned)
    if parsed is None:
        errors.append("Hemolysis must be a whole number from 0 to 6")
        return None
    try:
        return sample_service._normalize_hemolysis(parsed)
    except sample_service.SampleError as exc:
        errors.append(str(exc))
        return None


def _parse_optional_volume(value: str | None, errors: list[str]) -> float | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    parsed = _parse_float(cleaned)
    if parsed is None:
        errors.append("Volume must be numeric")
        return None
    if parsed < 0:
        errors.append("Volume must be zero or greater")
        return None
    return parsed


def _parse_optional_datetime(value: str | None, label: str, errors: list[str]) -> datetime | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    parsed = _parse_datetime(cleaned)
    if parsed is None:
        errors.append(f"{label} must use {DATE_FORMAT}")
    return parsed


def _display_optional(value: str | None) -> str | None:
    cleaned = _clean_text(value)
    return cleaned


def _display_collection(value: datetime | None) -> str | None:
    return value.strftime(DATE_FORMAT) if value else None


def _coerce_row_number(value, fallback: int) -> int:
    if isinstance(value, int):
        return value
    parsed = _parse_int(_clean_text(value))
    return parsed or fallback


def _unique_sample_ids(sample_ids: list[int]) -> list[int]:
    unique: list[int] = []
    for value in sample_ids:
        parsed = int(value)
        if parsed > 0 and parsed not in unique:
            unique.append(parsed)
    return unique


def _duplicate_text_values(values: list[str | None]) -> set[str]:
    counts: dict[str, int] = {}
    for value in values:
        cleaned = _clean_text(value)
        if not cleaned:
            continue
        counts[cleaned] = counts.get(cleaned, 0) + 1
    return {value for value, count in counts.items() if count > 1}


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_int(value: str | None) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _parse_float(value: str | None) -> float | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_datetime(value: str | None) -> datetime | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return datetime.strptime(cleaned, DATE_FORMAT)
    except ValueError:
        try:
            return datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        except ValueError:
            return None


def _volumes_match(current_volume: str, expected_volume: float | None) -> bool:
    parsed = _parse_float(current_volume)
    if parsed is None and expected_volume is None:
        return True
    if parsed is None or expected_volume is None:
        return False
    return abs(parsed - expected_volume) < 1e-9


def _xlsx_value_to_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_text(value)
